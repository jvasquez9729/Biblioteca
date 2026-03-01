#!/usr/bin/env python3
"""
Script para agregar fórmulas y datos de ejemplo al archivo Excel
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

def agregar_formulas_eerr(wb):
    """Agrega fórmulas al Estado de Resultados"""
    ws = wb["EERR"]

    # Fórmulas para año 2022 (columna B)
    # Utilidad Bruta = Ventas - Costo de Ventas
    ws['B7'] = "=B5-B6"
    ws['B7'].font = Font(color="000000", bold=True)

    # EBITDA = Utilidad Bruta - Gastos Admin - Gastos Ventas
    ws['B13'] = "=B7-B9-B10"
    ws['B13'].font = Font(color="000000", bold=True)

    # EBIT = EBITDA - Depreciación - Amortización
    ws['B14'] = "=B13-B11-B12"
    ws['B14'].font = Font(color="000000", bold=True)

    # UAI = EBIT + Otros Ingresos - Otros Gastos - Gastos Financieros
    ws['B19'] = "=B14+B16-B17-B18"
    ws['B19'].font = Font(color="000000", bold=True)

    # Impuesto = UAI * Tasa impuesto (referencia a configuración)
    ws['B21'] = "=B19*CONFIGURACIÓN!$B$15"
    ws['B21'].font = Font(color="008000")

    # Utilidad Neta = UAI - Impuesto
    ws['B22'] = "=B19-B21"
    ws['B22'].font = Font(color="000000", bold=True)

    # Análisis Vertical (AV%) para 2022 - todo como % de ventas
    for row in range(5, 23):
        ws.cell(row, 3).value = f"=B{row}/B$5"
        ws.cell(row, 3).number_format = "0.0%"

    # Copiar fórmulas para 2023 (columnas D, E, F)
    for row in range(7, 23):
        # Copiar fórmula de columna B a columna D
        if ws.cell(row, 2).value and isinstance(ws.cell(row, 2).value, str) and ws.cell(row, 2).value.startswith("="):
            formula = ws.cell(row, 2).value.replace("B", "D").replace("b", "d")
            ws.cell(row, 4).value = formula
            ws.cell(row, 4).font = Font(color="000000", bold=True)

    # AV% para 2023
    for row in range(5, 23):
        ws.cell(row, 5).value = f"=D{row}/D$5"
        ws.cell(row, 5).number_format = "0.0%"

    # AH% 2023 vs 2022
    for row in range(5, 23):
        ws.cell(row, 6).value = f"=(D{row}-B{row})/B{row}"
        ws.cell(row, 6).number_format = "0.0%"

    # Copiar fórmulas para 2024 (columnas G, H, I)
    for row in range(7, 23):
        if ws.cell(row, 2).value and isinstance(ws.cell(row, 2).value, str) and ws.cell(row, 2).value.startswith("="):
            formula = ws.cell(row, 2).value.replace("B", "G").replace("b", "g")
            ws.cell(row, 7).value = formula
            ws.cell(row, 7).font = Font(color="000000", bold=True)

    # AV% para 2024
    for row in range(5, 23):
        ws.cell(row, 8).value = f"=G{row}/G$5"
        ws.cell(row, 8).number_format = "0.0%"

    # AH% 2024 vs 2023
    for row in range(5, 23):
        ws.cell(row, 9).value = f"=(G{row}-D{row})/D{row}"
        ws.cell(row, 9).number_format = "0.0%"

    # Promedio
    for row in range(5, 23):
        ws.cell(row, 10).value = f"=AVERAGE(B{row},D{row},G{row})"

    # Formato de números
    for row in range(5, 23):
        for col in [2, 4, 7, 10]:  # Columnas B, D, G, J
            ws.cell(row, col).number_format = "#,##0"

    return ws

def agregar_formulas_esf(wb):
    """Agrega fórmulas al Balance General"""
    ws = wb["ESF"]

    # Total Activos Corrientes (fila 14)
    ws['B14'] = "=SUM(B8:B13)"
    ws['B14'].font = Font(color="000000", bold=True)
    ws['D14'] = "=SUM(D8:D13)"
    ws['D14'].font = Font(color="000000", bold=True)
    ws['G14'] = "=SUM(G8:G13)"
    ws['G14'].font = Font(color="000000", bold=True)

    # Total Activos No Corrientes (fila 21)
    ws['B21'] = "=SUM(B17:B20)"
    ws['B21'].font = Font(color="000000", bold=True)
    ws['D21'] = "=SUM(D17:D20)"
    ws['D21'].font = Font(color="000000", bold=True)
    ws['G21'] = "=SUM(G17:G20)"
    ws['G21'].font = Font(color="000000", bold=True)

    # Total Activos (fila 23)
    ws['B23'] = "=B14+B21"
    ws['B23'].font = Font(color="000000", bold=True, size=12)
    ws['D23'] = "=D14+D21"
    ws['D23'].font = Font(color="000000", bold=True, size=12)
    ws['G23'] = "=G14+G21"
    ws['G23'].font = Font(color="000000", bold=True, size=12)

    # Total Pasivos Corrientes (fila 37)
    ws['B37'] = "=SUM(B29:B36)"
    ws['B37'].font = Font(color="000000", bold=True)
    ws['D37'] = "=SUM(D29:D36)"
    ws['D37'].font = Font(color="000000", bold=True)
    ws['G37'] = "=SUM(G29:G36)"
    ws['G37'].font = Font(color="000000", bold=True)

    # Total Pasivos No Corrientes (fila 43)
    ws['B43'] = "=SUM(B40:B42)"
    ws['B43'].font = Font(color="000000", bold=True)
    ws['D43'] = "=SUM(D40:D42)"
    ws['D43'].font = Font(color="000000", bold=True)
    ws['G43'] = "=SUM(G40:G42)"
    ws['G43'].font = Font(color="000000", bold=True)

    # Total Pasivos (fila 45)
    ws['B45'] = "=B37+B43"
    ws['B45'].font = Font(color="000000", bold=True, size=12)
    ws['D45'] = "=D37+D43"
    ws['D45'].font = Font(color="000000", bold=True, size=12)
    ws['G45'] = "=G37+G43"
    ws['G45'].font = Font(color="000000", bold=True, size=12)

    # Total Patrimonio (fila 54)
    ws['B54'] = "=SUM(B49:B53)"
    ws['B54'].font = Font(color="000000", bold=True)
    ws['D54'] = "=SUM(D49:D53)"
    ws['D54'].font = Font(color="000000", bold=True)
    ws['G54'] = "=SUM(G49:G53)"
    ws['G54'].font = Font(color="000000", bold=True)

    # Total Pasivo + Patrimonio (fila 56)
    ws['B56'] = "=B45+B54"
    ws['B56'].font = Font(color="000000", bold=True, size=12)
    ws['D56'] = "=D45+D54"
    ws['D56'].font = Font(color="000000", bold=True, size=12)
    ws['G56'] = "=G45+G54"
    ws['G56'].font = Font(color="000000", bold=True, size=12)

    # Verificación de cuadre (fila 58)
    ws['B58'] = "=B23-B56"
    ws['B58'].font = Font(color="FF0000", bold=True)
    ws['D58'] = "=D23-D56"
    ws['D58'].font = Font(color="FF0000", bold=True)
    ws['G58'] = "=G23-G56"
    ws['G58'].font = Font(color="FF0000", bold=True)

    # Aplicar formato condicional (rojo si diferente de 0)
    for col in ['B', 'D', 'G']:
        ws[f'{col}58'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Análisis Vertical
    for row in range(8, 57):
        for col_pair in [(2, 3), (4, 5), (7, 8)]:  # (valor, AV%)
            ws.cell(row, col_pair[1]).value = f"={get_column_letter(col_pair[0])}{row}/{get_column_letter(col_pair[0])}$23"
            ws.cell(row, col_pair[1]).number_format = "0.0%"

    # Análisis Horizontal
    for row in range(8, 57):
        # 2023 vs 2022
        ws.cell(row, 6).value = f"=(D{row}-B{row})/B{row}"
        ws.cell(row, 6).number_format = "0.0%"
        # 2024 vs 2023
        ws.cell(row, 9).value = f"=(G{row}-D{row})/D{row}"
        ws.cell(row, 9).number_format = "0.0%"

    return ws

def agregar_formulas_indicadores(wb):
    """Agrega fórmulas a los Indicadores Financieros"""
    ws = wb["INDICADORES"]
    ws_esf = wb["ESF"]
    ws_eerr = wb["EERR"]

    # A. INDICADORES DE LIQUIDEZ (fila 7 en adelante)
    fila = 7

    # Razón Corriente = Activo Corriente / Pasivo Corriente
    ws[f'C{fila}'] = "=ESF!B14/ESF!B37"
    ws[f'D{fila}'] = "=ESF!D14/ESF!D37"
    ws[f'G{fila}'] = "=ESF!G14/ESF!G37"

    # Aplicar formato
    for col in ['C', 'D', 'G']:
        ws[f'{col}{fila}'].number_format = "0.00"

    # Variaciones
    ws[f'E{fila}'] = f"=D{fila}-C{fila}"
    ws[f'F{fila}'] = f"=(D{fila}-C{fila})/C{fila}"
    ws[f'H{fila}'] = f"=G{fila}-D{fila}"
    ws[f'I{fila}'] = f"=(G{fila}-D{fila})/D{fila}"

    ws[f'F{fila}'].number_format = "0.0%"
    ws[f'I{fila}'].number_format = "0.0%"

    # Semáforo (fila 7)
    ws[f'J{fila}'] = f'=IF(G{fila}>=1.5,"🟢 Óptimo",IF(G{fila}>=1,"🟡 Aceptable","🔴 Crítico"))'

    # Más indicadores se pueden agregar aquí...
    # Por brevidad, solo muestro la estructura

    return ws

def agregar_datos_ejemplo(wb):
    """Agrega datos de ejemplo para demostración"""

    # Configuración
    ws_config = wb["CONFIGURACIÓN"]
    ws_config['B5'] = "EMPRESA EJEMPLO S.A.S."
    ws_config['B7'] = "Manufactura"
    ws_config['B8'] = "1234"
    ws_config['B9'] = "COP"
    ws_config['B10'] = "Millones"
    ws_config['B11'] = 2024
    ws_config['B12'] = 3
    ws_config['B13'] = 5

    # Estado de Resultados - Datos de ejemplo
    ws_eerr = wb["EERR"]

    # Año 2022
    ws_eerr['B5'] = 15000  # Ventas
    ws_eerr['B6'] = 9000   # Costo de ventas
    ws_eerr['B9'] = 2500   # Gastos admin
    ws_eerr['B10'] = 1500  # Gastos ventas
    ws_eerr['B11'] = 400   # Depreciación
    ws_eerr['B12'] = 100   # Amortización
    ws_eerr['B16'] = 200   # Otros ingresos
    ws_eerr['B17'] = 150   # Otros gastos
    ws_eerr['B18'] = 300   # Gastos financieros

    # Año 2023
    ws_eerr['D5'] = 17500
    ws_eerr['D6'] = 10500
    ws_eerr['D9'] = 2700
    ws_eerr['D10'] = 1600
    ws_eerr['D11'] = 450
    ws_eerr['D12'] = 110
    ws_eerr['D16'] = 250
    ws_eerr['D17'] = 180
    ws_eerr['D18'] = 350

    # Año 2024
    ws_eerr['G5'] = 20000
    ws_eerr['G6'] = 12000
    ws_eerr['G9'] = 3000
    ws_eerr['G10'] = 1800
    ws_eerr['G11'] = 500
    ws_eerr['G12'] = 120
    ws_eerr['G16'] = 300
    ws_eerr['G17'] = 200
    ws_eerr['G18'] = 400

    # Balance General - Datos de ejemplo
    ws_esf = wb["ESF"]

    # Activos Corrientes 2022
    ws_esf['B8'] = 3000   # Efectivo
    ws_esf['B9'] = 2500   # CxC clientes
    ws_esf['B10'] = 500   # CxC socios
    ws_esf['B11'] = 2000  # Inventarios
    ws_esf['B12'] = 300   # Impuestos corrientes
    ws_esf['B13'] = 200   # Otros activos corrientes

    # Activos No Corrientes 2022
    ws_esf['B17'] = 8000  # PPE
    ws_esf['B18'] = 1000  # Propiedades inversión
    ws_esf['B19'] = 500   # Intangibles
    ws_esf['B20'] = 300   # Otros activos NC

    # Pasivos Corrientes 2022
    ws_esf['B29'] = 1500  # Oblig financieras CP
    ws_esf['B30'] = 1800  # Proveedores
    ws_esf['B31'] = 600   # Otras CxP
    ws_esf['B32'] = 200   # Provisiones
    ws_esf['B33'] = 0     # Dividendos
    ws_esf['B34'] = 150   # Impuestos por pagar
    ws_esf['B35'] = 400   # Beneficios empleados
    ws_esf['B36'] = 100   # Otros pasivos corrientes

    # Pasivos No Corrientes 2022
    ws_esf['B40'] = 3000  # Oblig financieras LP
    ws_esf['B41'] = 500   # CxP LP
    ws_esf['B42'] = 200   # Otros pasivos NC

    # Patrimonio 2022
    ws_esf['B49'] = 5000  # Capital
    ws_esf['B50'] = 800   # Reservas
    ws_esf['B51'] = 1250  # Resultado ejercicio
    ws_esf['B52'] = 2000  # Resultados acumulados
    ws_esf['B53'] = 500   # Superávit

    # Copiar estructura para 2023 y 2024 con incrementos
    for row in range(8, 54):
        if ws_esf.cell(row, 2).value and isinstance(ws_esf.cell(row, 2).value, (int, float)):
            ws_esf.cell(row, 4).value = ws_esf.cell(row, 2).value * 1.1  # +10%
            ws_esf.cell(row, 7).value = ws_esf.cell(row, 4).value * 1.12  # +12%

    return wb

def main():
    """Función principal"""
    print("Agregando fórmulas y datos de ejemplo...")

    archivo = "/home/user/Biblioteca/ControllerFinanciero/ControllerFinanciero_v1.0.xlsx"

    # Cargar workbook
    wb = openpyxl.load_workbook(archivo)

    print("Agregando fórmulas a Estado de Resultados...")
    agregar_formulas_eerr(wb)

    print("Agregando fórmulas a Balance General...")
    agregar_formulas_esf(wb)

    print("Agregando fórmulas a Indicadores...")
    agregar_formulas_indicadores(wb)

    print("Cargando datos de ejemplo...")
    agregar_datos_ejemplo(wb)

    # Guardar
    wb.save(archivo)
    print(f"\n✓ Fórmulas y datos agregados exitosamente a: {archivo}")

    return archivo

if __name__ == "__main__":
    main()
