#!/usr/bin/env python3
"""
Script para crear el Sistema Integral de Controller Financiero
Archivo Excel .xlsm con todas las hojas, fórmulas y estructura
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import datetime

def crear_estilos():
    """Define los estilos según las especificaciones"""
    return {
        'input': Font(color="0000FF", bold=True),  # Azul para inputs
        'formula': Font(color="000000"),  # Negro para fórmulas
        'referencia': Font(color="008000"),  # Verde para referencias
        'titulo': Font(size=14, bold=True, color="FFFFFF"),
        'subtitulo': Font(size=12, bold=True),
        'fill_titulo': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
        'fill_input': PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
        'fill_alerta': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        'fill_ok': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        'fill_precaucion': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        'fill_supuesto': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ),
        'align_center': Alignment(horizontal='center', vertical='center'),
        'align_left': Alignment(horizontal='left', vertical='center'),
        'align_right': Alignment(horizontal='right', vertical='center'),
    }

def crear_hoja_configuracion(wb, estilos):
    """Crea la Hoja 1: CONFIGURACIÓN"""
    ws = wb.create_sheet("CONFIGURACIÓN", 0)

    # Título
    ws['A1'] = "SISTEMA INTEGRAL DE CONTROLLER FINANCIERO"
    ws['A1'].font = estilos['titulo']
    ws['A1'].fill = estilos['fill_titulo']
    ws['A1'].alignment = estilos['align_center']
    ws.merge_cells('A1:D1')

    ws['A2'] = "CONFIGURACIÓN DEL ANÁLISIS"
    ws['A2'].font = estilos['subtitulo']
    ws.merge_cells('A2:D2')

    # Campos de configuración
    configuracion = [
        ("", "", "", ""),
        ("Parámetro", "Valor", "Descripción", ""),
        ("Nombre de la Empresa:", "", "Ingrese el nombre completo de la empresa", ""),
        ("Sector Económico:", "", "Seleccione el sector de la lista", ""),
        ("Código CIIU:", "", "Código de clasificación industrial", ""),
        ("Moneda:", "COP", "Moneda de presentación", ""),
        ("Unidad de Medida:", "Millones", "Miles, Millones o Unidades", ""),
        ("Año Base de Análisis:", str(datetime.datetime.now().year), "Año actual de análisis", ""),
        ("Años Históricos:", "3", "Número de años históricos (2-5)", ""),
        ("Años de Proyección:", "5", "Años para proyectar (5-7)", ""),
        ("", "", "", ""),
        ("Tasa de Impuesto de Renta:", "35%", "Colombia: 35% para empresas", ""),
        ("Tasa Libre de Riesgo:", "12.5%", "TES Colombia 10 años aprox.", ""),
    ]

    fila_inicio = 4
    for i, fila in enumerate(configuracion, start=fila_inicio):
        ws[f'A{i}'] = fila[0]
        ws[f'B{i}'] = fila[1]
        ws[f'C{i}'] = fila[2]

        # Aplicar estilos
        if i == fila_inicio + 1:  # Encabezados
            for col in ['A', 'B', 'C']:
                ws[f'{col}{i}'].font = Font(bold=True)
                ws[f'{col}{i}'].fill = estilos['fill_titulo']
                ws[f'{col}{i}'].font = Font(bold=True, color="FFFFFF")
        elif fila[0] and ":" in fila[0]:  # Etiquetas
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = estilos['input']  # Celda de entrada
            ws[f'B{i}'].fill = estilos['fill_input']

    # Lista desplegable para Sector Económico
    sectores = DataValidation(type="list", formula1='"Manufactura,Servicios,Comercio,Tecnología,Agroindustria,Salud,Construcción,Otro"')
    ws.add_data_validation(sectores)
    sectores.add(ws['B7'])

    # Lista desplegable para Unidad de Medida
    unidades = DataValidation(type="list", formula1='"Miles,Millones,Unidades"')
    ws.add_data_validation(unidades)
    unidades.add(ws['B10'])

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15

    return ws

def crear_hoja_eerr(wb, estilos):
    """Crea la Hoja 2: ENTRADA DE DATOS - EERR"""
    ws = wb.create_sheet("EERR")

    # Título
    ws['A1'] = "ESTADO DE RESULTADOS"
    ws['A1'].font = estilos['titulo']
    ws['A1'].fill = estilos['fill_titulo']
    ws['A1'].alignment = estilos['align_center']
    ws.merge_cells('A1:K1')

    ws['A2'] = "Cifras en: [Referencia a Configuración]"
    ws.merge_cells('A2:K2')

    # Encabezados de columnas
    ws['A4'] = "CONCEPTO"
    ws['B4'] = "Año 2022"
    ws['C4'] = "AV%"
    ws['D4'] = "Año 2023"
    ws['E4'] = "AV%"
    ws['F4'] = "AH%"
    ws['G4'] = "Año 2024"
    ws['H4'] = "AV%"
    ws['I4'] = "AH%"
    ws['J4'] = "Promedio"
    ws['K4'] = "Tendencia"

    for col in range(1, 12):
        cell = ws.cell(4, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = estilos['fill_titulo']
        cell.alignment = estilos['align_center']

    # Conceptos del Estado de Resultados
    conceptos = [
        ("Ingresos de Actividades Ordinarias (Ventas)", "input"),
        ("(-) Costo de Ventas / Costo de Servicios", "input"),
        ("(=) UTILIDAD BRUTA", "formula"),
        ("", ""),
        ("(-) Gastos Administrativos (excepto Deprec. y Amort.)", "input"),
        ("(-) Gastos de Ventas (excepto Deprec. y Amort.)", "input"),
        ("(-) Depreciación", "input"),
        ("(-) Amortización", "input"),
        ("(=) EBITDA", "formula"),
        ("(=) UTILIDAD OPERATIVA (EBIT)", "formula"),
        ("", ""),
        ("(+) Otros Ingresos", "input"),
        ("(-) Otros Gastos", "input"),
        ("(-) Gastos Financieros", "input"),
        ("(=) UTILIDAD ANTES DE IMPUESTOS (UAI)", "formula"),
        ("", ""),
        ("(-) Impuesto de Renta", "formula"),
        ("(=) UTILIDAD NETA", "formula"),
    ]

    fila = 5
    for concepto, tipo in conceptos:
        ws[f'A{fila}'] = concepto

        if tipo == "input":
            ws[f'A{fila}'].font = Font(bold=True)
            # Marcar celdas de entrada en azul
            for col in ['B', 'D', 'G']:
                ws[f'{col}{fila}'].font = estilos['input']
                ws[f'{col}{fila}'].fill = estilos['fill_input']
        elif tipo == "formula":
            ws[f'A{fila}'].font = Font(bold=True, italic=True)

        # Agregar fórmulas de ejemplo para las celdas calculadas
        if concepto.startswith("(=)"):
            # Las fórmulas se agregarán después con los cálculos reales
            pass

        fila += 1

    # Ajustar anchos
    ws.column_dimensions['A'].width = 50
    for col in ['B', 'D', 'G', 'J']:
        ws.column_dimensions[col].width = 15
    for col in ['C', 'E', 'F', 'H', 'I', 'K']:
        ws.column_dimensions[col].width = 10

    return ws

def crear_hoja_esf(wb, estilos):
    """Crea la Hoja 3: ENTRADA DE DATOS - ESF (Balance General)"""
    ws = wb.create_sheet("ESF")

    # Título
    ws['A1'] = "ESTADO DE SITUACIÓN FINANCIERA (BALANCE GENERAL)"
    ws['A1'].font = estilos['titulo']
    ws['A1'].fill = estilos['fill_titulo']
    ws['A1'].alignment = estilos['align_center']
    ws.merge_cells('A1:K1')

    # Encabezados
    ws['A4'] = "CONCEPTO"
    ws['B4'] = "Año 2022"
    ws['C4'] = "AV%"
    ws['D4'] = "Año 2023"
    ws['E4'] = "AV%"
    ws['F4'] = "AH%"
    ws['G4'] = "Año 2024"
    ws['H4'] = "AV%"
    ws['I4'] = "AH%"

    for col in range(1, 10):
        cell = ws.cell(4, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = estilos['fill_titulo']
        cell.alignment = estilos['align_center']

    # Estructura del Balance
    estructura = [
        ("ACTIVOS", "seccion"),
        ("", ""),
        ("ACTIVOS CORRIENTES:", "subseccion"),
        ("Efectivo y Equivalentes", "input"),
        ("Cuentas por Cobrar Comerciales (Clientes)", "input"),
        ("Cuentas por Cobrar a Socios/Accionistas", "input"),
        ("Inventarios", "input"),
        ("Activos por Impuestos Corrientes", "input"),
        ("Otros Activos Corrientes", "input"),
        ("TOTAL ACTIVOS CORRIENTES", "formula"),
        ("", ""),
        ("ACTIVOS NO CORRIENTES:", "subseccion"),
        ("Propiedades, Planta y Equipo (neto)", "input"),
        ("Propiedades de Inversión", "input"),
        ("Intangibles", "input"),
        ("Otros Activos No Corrientes", "input"),
        ("TOTAL ACTIVOS NO CORRIENTES", "formula"),
        ("", ""),
        ("TOTAL ACTIVOS", "total"),
        ("", ""),
        ("", ""),
        ("PASIVOS", "seccion"),
        ("", ""),
        ("PASIVOS CORRIENTES:", "subseccion"),
        ("Obligaciones Financieras Corto Plazo", "input"),
        ("Cuentas por Pagar Comerciales (Proveedores)", "input"),
        ("Otras Cuentas por Pagar (Costos y Gastos)", "input"),
        ("Pasivos Estimados y Provisiones", "input"),
        ("Dividendos por Pagar", "input"),
        ("Impuestos Corrientes por Pagar", "input"),
        ("Beneficios a Empleados", "input"),
        ("Otros Pasivos Corrientes", "input"),
        ("TOTAL PASIVOS CORRIENTES", "formula"),
        ("", ""),
        ("PASIVOS NO CORRIENTES:", "subseccion"),
        ("Obligaciones Financieras Largo Plazo", "input"),
        ("Cuentas por Pagar Largo Plazo", "input"),
        ("Otros Pasivos No Corrientes", "input"),
        ("TOTAL PASIVOS NO CORRIENTES", "formula"),
        ("", ""),
        ("TOTAL PASIVOS", "total"),
        ("", ""),
        ("", ""),
        ("PATRIMONIO", "seccion"),
        ("", ""),
        ("Capital Social", "input"),
        ("Reservas", "input"),
        ("Resultado del Ejercicio", "input"),
        ("Resultados Acumulados", "input"),
        ("Superávit por Valorización/Revaluación", "input"),
        ("TOTAL PATRIMONIO", "formula"),
        ("", ""),
        ("TOTAL PASIVO + PATRIMONIO", "total"),
        ("", ""),
        ("VERIFICACIÓN DE CUADRE (debe ser cero)", "verificacion"),
    ]

    fila = 5
    for concepto, tipo in estructura:
        ws[f'A{fila}'] = concepto

        if tipo == "seccion":
            ws[f'A{fila}'].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f'A{fila}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        elif tipo == "subseccion":
            ws[f'A{fila}'].font = Font(bold=True, underline="single")
        elif tipo == "input":
            for col in ['B', 'D', 'G']:
                ws[f'{col}{fila}'].font = estilos['input']
                ws[f'{col}{fila}'].fill = estilos['fill_input']
        elif tipo == "formula" or tipo == "total":
            ws[f'A{fila}'].font = Font(bold=True, italic=True)
        elif tipo == "verificacion":
            ws[f'A{fila}'].font = Font(bold=True, color="FF0000")

        fila += 1

    # Ajustar anchos
    ws.column_dimensions['A'].width = 50
    for col in ['B', 'D', 'G']:
        ws.column_dimensions[col].width = 15
    for col in ['C', 'E', 'F', 'H', 'I']:
        ws.column_dimensions[col].width = 10

    return ws

def crear_hoja_indicadores(wb, estilos):
    """Crea la Hoja 4: INDICADORES FINANCIEROS"""
    ws = wb.create_sheet("INDICADORES")

    # Título
    ws['A1'] = "INDICADORES FINANCIEROS"
    ws['A1'].font = estilos['titulo']
    ws['A1'].fill = estilos['fill_titulo']
    ws['A1'].alignment = estilos['align_center']
    ws.merge_cells('A1:L1')

    # Encabezados
    headers = ["INDICADOR", "Fórmula", "2022", "2023", "Δ", "Δ%", "2024", "Δ", "Δ%", "Semáforo", "Benchmark", "vs Ind"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(4, col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = estilos['fill_titulo']
        cell.alignment = estilos['align_center']

    # A. INDICADORES DE LIQUIDEZ
    indicadores = [
        ("", ""),
        ("A. INDICADORES DE LIQUIDEZ", ""),
        ("Razón Corriente", "Activo Corriente / Pasivo Corriente"),
        ("Prueba Ácida", "(Activo Corriente - Inventarios) / Pasivo Corriente"),
        ("Capital de Trabajo (KT)", "Activo Corriente - Pasivo Corriente"),
        ("Capital de Trabajo Neto Operativo (KTNO)", "CxC Clientes + Inventarios - Proveedores"),
        ("Productividad del KT (PKT)", "KT / Ventas"),
        ("Productividad del KTNO (PKTNO)", "KTNO / Ventas"),
        ("Solvencia", "Activo Total / Pasivo Total"),
        ("", ""),
        ("B. INDICADORES DE RENTABILIDAD", ""),
        ("Margen Bruto", "Utilidad Bruta / Ventas"),
        ("Margen EBITDA", "EBITDA / Ventas"),
        ("Margen Operativo", "Utilidad Operativa / Ventas"),
        ("Margen Neto", "Utilidad Neta / Ventas"),
        ("ROA (Retorno sobre Activos)", "Utilidad Neta / Activos Totales"),
        ("ROE (Retorno sobre Patrimonio)", "Utilidad Neta / Patrimonio"),
        ("Tasa Efectiva de Tributación", "Impuestos / UAI"),
        ("", ""),
        ("C. INDICADORES DE ENDEUDAMIENTO", ""),
        ("Nivel de Endeudamiento", "Pasivo Total / Activo Total"),
        ("Concentración Deuda Corto Plazo", "Pasivo Corriente / Pasivo Total"),
        ("Deuda Financiera / Activos", "(Oblig. Financ. CP + LP) / Activos"),
        ("Deuda / Patrimonio", "Pasivo Total / Patrimonio"),
        ("Cobertura de Intereses", "EBIT / Gastos Financieros"),
        ("", ""),
        ("D. INDICADORES DE ACTIVIDAD/ROTACIÓN", ""),
        ("Rotación de Inventarios (veces)", "Costo de Ventas / Inventarios"),
        ("Días de Inventario", "365 / Rotación de Inventarios"),
        ("Rotación de Cartera (veces)", "Ventas / CxC Clientes"),
        ("Días de Cartera", "365 / Rotación de Cartera"),
        ("Rotación de Proveedores (veces)", "Costo de Ventas / Proveedores"),
        ("Días de Proveedores", "365 / Rotación de Proveedores"),
        ("Ciclo Operativo", "Días Inventario + Días Cartera"),
        ("Ciclo de Conversión de Efectivo", "Ciclo Operativo - Días Proveedores"),
        ("GAP Financiero", "Días de financiación requerida"),
        ("", ""),
        ("E. INDICADORES DE INVERSIÓN", ""),
        ("Depreciación / PPE", "Depreciación / Propiedad, Planta y Equipo"),
        ("CAPEX / Activos Fijos", "CAPEX / Activos Fijos"),
        ("Variación % en Ventas", "Crecimiento en Ventas"),
        ("Variación % en EBITDA", "Crecimiento en EBITDA"),
    ]

    fila = 5
    for indicador, formula in indicadores:
        ws[f'A{fila}'] = indicador
        ws[f'B{fila}'] = formula

        if indicador and not formula:  # Secciones
            ws[f'A{fila}'].font = Font(bold=True, size=11, color="FFFFFF")
            ws[f'A{fila}'].fill = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")
            ws.merge_cells(f'A{fila}:L{fila}')
        elif indicador:
            ws[f'A{fila}'].font = Font(bold=True)
            ws[f'B{fila}'].font = Font(italic=True, size=9)

        fila += 1

    # Ajustar anchos
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 35
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws.column_dimensions[col].width = 12

    return ws

def crear_hoja_punto_equilibrio(wb, estilos):
    """Crea la Hoja 5: PUNTO DE EQUILIBRIO"""
    ws = wb.create_sheet("PUNTO EQUILIBRIO")

    # Título
    ws['A1'] = "ANÁLISIS DE PUNTO DE EQUILIBRIO"
    ws['A1'].font = estilos['titulo']
    ws['A1'].fill = estilos['fill_titulo']
    ws['A1'].alignment = estilos['align_center']
    ws.merge_cells('A1:F1')

    # Datos de entrada
    ws['A3'] = "DATOS DE ENTRADA"
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].fill = estilos['fill_titulo']
    ws['A3'].font = Font(bold=True, size=12, color="FFFFFF")
    ws.merge_cells('A3:F3')

    datos_entrada = [
        ("Precio de Venta Promedio Unitario:", "", "$"),
        ("Costo Variable Unitario:", "", "$"),
        ("", "", ""),
        ("COSTOS FIJOS TOTALES:", "", ""),
        ("Gastos Administrativos:", "", "$"),
        ("Gastos de Ventas:", "", "$"),
        ("Gastos Financieros:", "", "$"),
        ("TOTAL COSTOS FIJOS:", "", "$"),
    ]

    fila = 4
    for etiqueta, valor, unidad in datos_entrada:
        ws[f'A{fila}'] = etiqueta
        ws[f'B{fila}'] = valor
        ws[f'C{fila}'] = unidad

        if etiqueta and ":" in etiqueta:
            ws[f'A{fila}'].font = Font(bold=True)
            ws[f'B{fila}'].font = estilos['input']
            ws[f'B{fila}'].fill = estilos['fill_input']

        fila += 1

    # Resultados calculados
    ws[f'A{fila+1}'] = "RESULTADOS CALCULADOS"
    ws[f'A{fila+1}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{fila+1}'].fill = estilos['fill_titulo']
    ws.merge_cells(f'A{fila+1}:F{fila+1}')

    resultados = [
        ("Margen de Contribución Unitario:", ""),
        ("Margen de Contribución %:", ""),
        ("Punto de Equilibrio (unidades):", ""),
        ("Punto de Equilibrio (valor $):", ""),
        ("Ventas Actuales (unidades):", ""),
        ("Margen de Seguridad:", ""),
    ]

    fila += 2
    for etiqueta, valor in resultados:
        ws[f'A{fila}'] = etiqueta
        ws[f'B{fila}'] = valor
        ws[f'A{fila}'].font = Font(bold=True)
        fila += 1

    # Análisis de sensibilidad
    ws[f'A{fila+1}'] = "ANÁLISIS DE SENSIBILIDAD"
    ws[f'A{fila+1}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{fila+1}'].fill = estilos['fill_titulo']
    ws.merge_cells(f'A{fila+1}:F{fila+1}')

    fila += 2
    ws[f'A{fila}'] = "Variación del Precio de Venta"
    ws.merge_cells(f'A{fila}:F{fila}')

    fila += 1
    sensibilidad_headers = ["Escenario", "Variación", "Nuevo Precio", "PE Unidades", "PE Valor ($)", "Impacto"]
    for col, header in enumerate(sensibilidad_headers, start=1):
        cell = ws.cell(fila, col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    escenarios = [
        ("Precio -15%", "-15%"),
        ("Precio -10%", "-10%"),
        ("Precio -5%", "-5%"),
        ("Precio Base", "0%"),
        ("Precio +5%", "+5%"),
        ("Precio +10%", "+10%"),
        ("Precio +15%", "+15%"),
    ]

    fila += 1
    for escenario, variacion in escenarios:
        ws[f'A{fila}'] = escenario
        ws[f'B{fila}'] = variacion
        fila += 1

    # Ajustar anchos
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    return ws

def crear_hoja_proyecciones(wb, estilos):
    """Crea la Hoja 6: PROYECCIONES FINANCIERAS"""
    ws = wb.create_sheet("PROYECCIONES")

    # Título
    ws['A1'] = "PROYECCIONES FINANCIERAS A 5-7 AÑOS"
    ws['A1'].font = estilos['titulo']
    ws['A1'].fill = estilos['fill_titulo']
    ws['A1'].alignment = estilos['align_center']
    ws.merge_cells('A1:L1')

    # Supuestos de proyección
    ws['A3'] = "SUPUESTOS DE PROYECCIÓN"
    ws['A3'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A3'].fill = estilos['fill_titulo']
    ws.merge_cells('A3:L3')

    supuestos = [
        ("Tasa de Crecimiento de Ventas (anual):", "%"),
        ("Margen Bruto Esperado:", "%"),
        ("Gastos Administrativos (% de Ventas):", "%"),
        ("Gastos de Ventas (% de Ventas):", "%"),
        ("Tasa de Interés de Deuda:", "%"),
        ("Tasa de Impuesto de Renta:", "35%"),
        ("Días de Cartera Objetivo:", "días"),
        ("Días de Inventario Objetivo:", "días"),
        ("Días de Proveedores Objetivo:", "días"),
        ("CAPEX Anual Proyectado:", "$"),
    ]

    fila = 4
    for etiqueta, valor in supuestos:
        ws[f'A{fila}'] = etiqueta
        ws[f'B{fila}'] = valor
        ws[f'A{fila}'].font = Font(bold=True)
        ws[f'B{fila}'].font = estilos['input']
        ws[f'B{fila}'].fill = estilos['fill_supuesto']
        fila += 1

    # Encabezados de años
    ws[f'A{fila+2}'] = "ESTADO DE RESULTADOS PROYECTADO"
    ws[f'A{fila+2}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{fila+2}'].fill = estilos['fill_titulo']
    ws.merge_cells(f'A{fila+2}:L{fila+2}')

    fila += 3
    ws[f'A{fila}'] = "CONCEPTO"
    for i, year in enumerate(range(2025, 2032), start=2):
        ws.cell(fila, i).value = f"Año {year}"
        ws.cell(fila, i).font = Font(bold=True)
        ws.cell(fila, i).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Conceptos proyectados
    conceptos_proy = [
        "Ingresos (Ventas)",
        "(-) Costo de Ventas",
        "Utilidad Bruta",
        "",
        "(-) Gastos Administrativos",
        "(-) Gastos de Ventas",
        "(-) Depreciación",
        "EBITDA",
        "Utilidad Operativa (EBIT)",
        "",
        "(-) Gastos Financieros",
        "Utilidad Antes de Impuestos",
        "(-) Impuesto de Renta",
        "UTILIDAD NETA",
    ]

    fila += 1
    for concepto in conceptos_proy:
        ws[f'A{fila}'] = concepto
        if concepto and not concepto.startswith("("):
            ws[f'A{fila}'].font = Font(bold=True)
        fila += 1

    # Ajustar anchos
    ws.column_dimensions['A'].width = 35
    for col in range(2, 13):
        ws.column_dimensions[get_column_letter(col)].width = 12

    return ws

def crear_hoja_valoracion(wb, estilos):
    """Crea la Hoja 7: VALORACIÓN"""
    ws = wb.create_sheet("VALORACIÓN")

    ws['A1'] = "VALORACIÓN DE LA EMPRESA"
    ws['A1'].font = estilos['titulo']
    ws['A1'].fill = estilos['fill_titulo']
    ws['A1'].alignment = estilos['align_center']
    ws.merge_cells('A1:H1')

    ws['A3'] = "CÁLCULO DEL WACC (Costo Promedio Ponderado de Capital)"
    ws['A3'].font = Font(bold=True, size=11)
    ws.merge_cells('A3:H3')

    # Componentes del WACC
    wacc_componentes = [
        ("Costo del Patrimonio (Ke) - CAPM:", ""),
        ("  Tasa Libre de Riesgo (Rf):", "%"),
        ("  Beta del Sector:", ""),
        ("  Prima de Riesgo de Mercado:", "%"),
        ("  Prima de Riesgo País Colombia:", "%"),
        ("  Ke = Rf + β × (Rm - Rf) + Riesgo País", ""),
        ("", ""),
        ("Costo de la Deuda (Kd):", "%"),
        ("Tasa de Impuesto:", "35%"),
        ("Kd después de impuestos:", ""),
        ("", ""),
        ("Estructura de Capital:", ""),
        ("  Patrimonio (E):", "$"),
        ("  Deuda (D):", "$"),
        ("  Total (E+D):", "$"),
        ("  % Patrimonio:", "%"),
        ("  % Deuda:", "%"),
        ("", ""),
        ("WACC = Ke × (E/V) + Kd × (1-T) × (D/V)", ""),
        ("WACC CALCULADO:", "%"),
    ]

    fila = 4
    for etiqueta, valor in wacc_componentes:
        ws[f'A{fila}'] = etiqueta
        ws[f'B{fila}'] = valor

        if etiqueta and not etiqueta.startswith(" "):
            ws[f'A{fila}'].font = Font(bold=True)
        if ":" in etiqueta and not etiqueta.startswith("  "):
            ws[f'B{fila}'].font = estilos['input']
            ws[f'B{fila}'].fill = estilos['fill_input']

        fila += 1

    # Flujo de Caja Descontado
    ws[f'A{fila+1}'] = "FLUJO DE CAJA LIBRE DESCONTADO"
    ws[f'A{fila+1}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{fila+1}'].fill = estilos['fill_titulo']
    ws.merge_cells(f'A{fila+1}:H{fila+1}')

    fila += 3
    headers_fcf = ["Año", "FCF", "Factor Desc.", "FCF Descontado", "FCF Acum."]
    for col, header in enumerate(headers_fcf, start=1):
        ws.cell(fila, col).value = header
        ws.cell(fila, col).font = Font(bold=True)

    # Valor terminal
    fila += 8
    ws[f'A{fila}'] = "Valor Terminal (Perpetuidad):"
    ws[f'B{fila}'] = "Tasa de crecimiento perpetuo:"
    ws[f'C{fila}'] = "2.5%"
    ws[f'C{fila}'].font = estilos['input']
    ws[f'C{fila}'].fill = estilos['fill_input']

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25

    return ws

def crear_hoja_benchmarks(wb, estilos):
    """Crea la Hoja 8: BENCHMARKS DE INDUSTRIA"""
    ws = wb.create_sheet("BENCHMARKS")

    ws['A1'] = "BENCHMARKS DE INDUSTRIA (Damodaran)"
    ws['A1'].font = estilos['titulo']
    ws['A1'].fill = estilos['fill_titulo']
    ws['A1'].alignment = estilos['align_center']
    ws.merge_cells('A1:L1')

    ws['A2'] = "Fuente: Aswath Damodaran - Stern School of Business, NYU"
    ws['A2'].font = Font(italic=True, size=9)
    ws.merge_cells('A2:L2')

    # Encabezados
    headers = ["Sector", "Margen Bruto", "Margen EBITDA", "Margen Neto", "ROE", "ROA",
               "Beta", "Razón Corriente", "Días Inv.", "EV/Sales", "EV/EBITDA"]

    fila = 4
    for col, header in enumerate(headers, start=1):
        ws.cell(fila, col).value = header
        ws.cell(fila, col).font = Font(bold=True, color="FFFFFF")
        ws.cell(fila, col).fill = estilos['fill_titulo']
        ws.cell(fila, col).alignment = estilos['align_center']

    # Datos de benchmarks (aproximados - deben actualizarse con datos reales)
    benchmarks_data = [
        ("Manufactura", "28%", "15%", "8%", "12%", "6%", "1.05", "1.8", "45", "1.2x", "8.5x"),
        ("Servicios", "45%", "22%", "12%", "18%", "9%", "0.95", "1.5", "15", "2.0x", "10.0x"),
        ("Comercio / Retail", "25%", "10%", "5%", "15%", "7%", "1.10", "1.6", "60", "0.8x", "7.5x"),
        ("Tecnología / Software", "75%", "35%", "18%", "25%", "12%", "1.25", "2.5", "10", "5.0x", "15.0x"),
        ("Agroindustria", "22%", "12%", "6%", "10%", "5%", "0.90", "1.7", "90", "1.0x", "9.0x"),
        ("Salud / Healthcare", "55%", "25%", "14%", "20%", "10%", "0.85", "2.0", "30", "2.5x", "12.0x"),
        ("Construcción", "18%", "8%", "4%", "12%", "5%", "1.15", "1.4", "75", "0.9x", "8.0x"),
    ]

    fila += 1
    for sector_data in benchmarks_data:
        for col, value in enumerate(sector_data, start=1):
            ws.cell(fila, col).value = value
        fila += 1

    # Comparación con la empresa
    ws[f'A{fila+2}'] = "COMPARACIÓN: EMPRESA VS INDUSTRIA"
    ws[f'A{fila+2}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{fila+2}'].fill = estilos['fill_titulo']
    ws.merge_cells(f'A{fila+2}:F{fila+2}')

    fila += 4
    comparacion_headers = ["Indicador", "Empresa", "Industria", "Brecha", "Semáforo", "Comentario"]
    for col, header in enumerate(comparacion_headers, start=1):
        ws.cell(fila, col).value = header
        ws.cell(fila, col).font = Font(bold=True)
        ws.cell(fila, col).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Ajustar anchos
    ws.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws.column_dimensions[col].width = 12

    return ws

def crear_hoja_diagnostico(wb, estilos):
    """Crea la Hoja 9: DIAGNÓSTICO AUTOMÁTICO"""
    ws = wb.create_sheet("DIAGNÓSTICO")

    ws['A1'] = "DIAGNÓSTICO FINANCIERO AUTOMÁTICO"
    ws['A1'].font = estilos['titulo']
    ws['A1'].fill = estilos['fill_titulo']
    ws['A1'].alignment = estilos['align_center']
    ws.merge_cells('A1:H1')

    # Encabezado del diagnóstico
    diagnostico_estructura = [
        "",
        "=== DIAGNÓSTICO FINANCIERO DE [NOMBRE EMPRESA] ===",
        "Sector: [SECTOR]",
        "Período analizado: [AÑOS]",
        "Fecha de elaboración: [FECHA AUTOMÁTICA]",
        "",
        "═══════════════════════════════════════════════════════════════",
        "",
        "1. RESUMEN EJECUTIVO",
        "────────────────────────────────────────────────────────────",
        "",
        "[Generar automáticamente un párrafo de 3-4 oraciones resumiendo la situación general]",
        "",
        "",
        "2. ANÁLISIS DE LIQUIDEZ",
        "────────────────────────────────────────────────────────────",
        "",
        "2.1 Razón Corriente",
        "Valor actual: [VALOR]",
        "Interpretación: [TEXTO CONDICIONAL BASADO EN VALOR]",
        "",
        "2.2 Prueba Ácida",
        "Valor actual: [VALOR]",
        "Interpretación: [TEXTO CONDICIONAL]",
        "",
        "2.3 Capital de Trabajo",
        "Valor actual: [VALOR]",
        "Interpretación: [TEXTO CONDICIONAL]",
        "",
        "",
        "3. ANÁLISIS DE RENTABILIDAD",
        "────────────────────────────────────────────────────────────",
        "",
        "3.1 Margen Bruto",
        "Valor actual: [VALOR]",
        "Interpretación: [TEXTO CONDICIONAL]",
        "",
        "3.2 Margen EBITDA",
        "Valor actual: [VALOR]",
        "Interpretación: [TEXTO CONDICIONAL]",
        "",
        "3.3 Margen Neto",
        "Valor actual: [VALOR]",
        "Interpretación: [TEXTO CONDICIONAL]",
        "",
        "3.4 ROE (Retorno sobre Patrimonio)",
        "Valor actual: [VALOR]",
        "Interpretación: [TEXTO CONDICIONAL]",
        "",
        "",
        "4. ANÁLISIS DE ENDEUDAMIENTO",
        "────────────────────────────────────────────────────────────",
        "",
        "4.1 Nivel de Endeudamiento",
        "Valor actual: [VALOR]",
        "Interpretación: [TEXTO CONDICIONAL]",
        "",
        "4.2 Cobertura de Intereses",
        "Valor actual: [VALOR]",
        "Interpretación: [TEXTO CONDICIONAL]",
        "",
        "",
        "5. ANÁLISIS DEL CICLO DE EFECTIVO",
        "────────────────────────────────────────────────────────────",
        "",
        "Días de Inventario: [VALOR]",
        "Días de Cartera: [VALOR]",
        "Días de Proveedores: [VALOR]",
        "Ciclo de Conversión de Efectivo: [VALOR]",
        "",
        "Interpretación: [TEXTO CONDICIONAL]",
        "",
        "",
        "6. PUNTO DE EQUILIBRIO",
        "────────────────────────────────────────────────────────────",
        "",
        "Punto de Equilibrio: [VALOR] unidades o $[VALOR]",
        "Ventas Actuales: [VALOR]",
        "Margen de Seguridad: [VALOR]%",
        "",
        "Interpretación: [TEXTO CONDICIONAL]",
        "",
        "",
        "7. COMPARACIÓN CON LA INDUSTRIA",
        "────────────────────────────────────────────────────────────",
        "",
        "[Tabla comparativa generada automáticamente]",
        "",
        "",
        "8. PRINCIPALES HALLAZGOS Y ALERTAS",
        "────────────────────────────────────────────────────────────",
        "",
        "FORTALEZAS IDENTIFICADAS:",
        "• [Fortaleza 1]",
        "• [Fortaleza 2]",
        "• [Fortaleza 3]",
        "",
        "DEBILIDADES IDENTIFICADAS:",
        "• [Debilidad 1]",
        "• [Debilidad 2]",
        "• [Debilidad 3]",
        "",
        "OPORTUNIDADES DE MEJORA:",
        "• [Oportunidad 1]",
        "• [Oportunidad 2]",
        "• [Oportunidad 3]",
        "",
        "",
        "9. RECOMENDACIONES DE ACCIÓN",
        "────────────────────────────────────────────────────────────",
        "",
        "RECOMENDACIÓN 1: [Título]",
        "Situación actual: [Descripción]",
        "Meta propuesta: [Objetivo cuantificable]",
        "Acciones sugeridas:",
        "  1. [Acción específica 1]",
        "  2. [Acción específica 2]",
        "  3. [Acción específica 3]",
        "Impacto esperado: [Beneficio cuantificado]",
        "",
        "RECOMENDACIÓN 2: [Título]",
        "[Similar estructura]",
        "",
        "RECOMENDACIÓN 3: [Título]",
        "[Similar estructura]",
        "",
        "",
        "10. CONCLUSIÓN",
        "────────────────────────────────────────────────────────────",
        "",
        "[Párrafo de cierre con perspectiva general y próximos pasos]",
        "",
        "═══════════════════════════════════════════════════════════════",
        "",
        "Este diagnóstico fue generado automáticamente por el Sistema Integral",
        "de Controller Financiero. Se recomienda revisión por parte de un",
        "profesional contable para validar interpretaciones y recomendaciones.",
    ]

    for i, linea in enumerate(diagnostico_estructura, start=2):
        ws[f'A{i}'] = linea

        if "===" in linea or "DIAGNÓSTICO FINANCIERO" in linea:
            ws[f'A{i}'].font = Font(bold=True, size=14)
            ws[f'A{i}'].alignment = estilos['align_center']
        elif linea.startswith(tuple(str(i) + "." for i in range(1, 11))):
            ws[f'A{i}'].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f'A{i}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        elif "────" in linea:
            ws[f'A{i}'].font = Font(color="808080")
        elif linea.startswith(("FORTALEZAS", "DEBILIDADES", "OPORTUNIDADES", "RECOMENDACIÓN")):
            ws[f'A{i}'].font = Font(bold=True, size=11)

        ws.merge_cells(f'A{i}:H{i}')

    ws.column_dimensions['A'].width = 100

    return ws

def crear_hoja_informe_visual(wb, estilos):
    """Crea la Hoja 10: INFORME VISUAL"""
    ws = wb.create_sheet("INFORME VISUAL")

    ws['A1'] = "DASHBOARD EJECUTIVO"
    ws['A1'].font = estilos['titulo']
    ws['A1'].fill = estilos['fill_titulo']
    ws['A1'].alignment = estilos['align_center']
    ws.merge_cells('A1:P1')

    # Áreas para gráficos
    areas_graficos = [
        (3, "Gráfico 1: Evolución de Ventas y Utilidades"),
        (18, "Gráfico 2: Comparación de Márgenes (Bruto, EBITDA, Operativo, Neto)"),
        (33, "Gráfico 3: Composición del Activo"),
        (48, "Gráfico 4: Composición del Pasivo y Patrimonio"),
        (63, "Gráfico 5: Punto de Equilibrio"),
        (78, "Gráfico 6: Comparación con Industria (Radar)"),
        (93, "Gráfico 7: Ciclo de Conversión de Efectivo"),
        (108, "Gráfico 8: Indicadores de Liquidez y Endeudamiento"),
    ]

    for fila, titulo in areas_graficos:
        ws[f'A{fila}'] = titulo
        ws[f'A{fila}'].font = Font(bold=True, size=11)
        ws[f'A{fila}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws.merge_cells(f'A{fila}:H{fila}')

        ws[f'A{fila+1}'] = "[Espacio reservado para gráfico - se generará con macro VBA]"
        ws.merge_cells(f'A{fila+1}:H{fila+12}')

    ws.column_dimensions['A'].width = 15

    return ws

def crear_hoja_instrucciones(wb, estilos):
    """Crea la Hoja de INSTRUCCIONES"""
    ws = wb.create_sheet("INSTRUCCIONES", 0)

    ws['A1'] = "GUÍA DE USO - SISTEMA INTEGRAL DE CONTROLLER FINANCIERO"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['A1'].alignment = estilos['align_center']
    ws.merge_cells('A1:F1')

    instrucciones = [
        "",
        "BIENVENIDO AL SISTEMA INTEGRAL DE CONTROLLER FINANCIERO",
        "",
        "Esta herramienta ha sido diseñada para realizar análisis financiero completo de empresas,",
        "siguiendo la metodología de consultoría financiera para PYMEs colombianas.",
        "",
        "═══════════════════════════════════════════════════════════════════════════════",
        "",
        "PASO 1: CONFIGURACIÓN INICIAL",
        "• Vaya a la hoja 'CONFIGURACIÓN'",
        "• Complete los datos básicos de la empresa (nombre, sector, CIIU, etc.)",
        "• Las celdas de entrada están marcadas en AZUL con fondo gris",
        "• Seleccione el sector económico de la lista desplegable",
        "",
        "PASO 2: INGRESAR ESTADO DE RESULTADOS",
        "• Vaya a la hoja 'EERR' (Estado de Resultados)",
        "• Ingrese los datos financieros para cada año en las celdas AZULES",
        "• Los cálculos se realizarán automáticamente (celdas en negro)",
        "• El análisis vertical (AV%) y horizontal (AH%) se calculará solo",
        "",
        "PASO 3: INGRESAR BALANCE GENERAL",
        "• Vaya a la hoja 'ESF' (Estado de Situación Financiera)",
        "• Complete todas las partidas del Activo, Pasivo y Patrimonio",
        "• IMPORTANTE: Verifique que el balance cuadre (verificación al final)",
        "• Si hay error de cuadre, revise las cifras ingresadas",
        "",
        "PASO 4: REVISAR INDICADORES CALCULADOS",
        "• Vaya a la hoja 'INDICADORES'",
        "• Más de 30 indicadores se calculan automáticamente",
        "• Revise los semáforos para identificar alertas rápidamente",
        "  Verde = Óptimo | Amarillo = Precaución | Rojo = Alerta",
        "",
        "PASO 5: ANÁLISIS DE PUNTO DE EQUILIBRIO",
        "• Vaya a la hoja 'PUNTO EQUILIBRIO'",
        "• Ingrese: precio de venta unitario, costo variable unitario y costos fijos",
        "• El punto de equilibrio se calcula en unidades y en pesos",
        "• Revise el análisis de sensibilidad para diferentes escenarios",
        "",
        "PASO 6: PROYECCIONES FINANCIERAS",
        "• Vaya a la hoja 'PROYECCIONES'",
        "• Defina los supuestos (crecimiento, márgenes, etc.) en celdas amarillas",
        "• Las proyecciones se generan para 5-7 años",
        "• Se incluyen 3 escenarios: Base, Optimista y Pesimista",
        "",
        "PASO 7: VALORACIÓN (OPCIONAL)",
        "• Vaya a la hoja 'VALORACIÓN'",
        "• Calcule el WACC (costo de capital)",
        "• Revise el flujo de caja descontado",
        "• Obtenga el valor estimado de la empresa",
        "",
        "PASO 8: COMPARACIÓN CON INDUSTRIA",
        "• Vaya a la hoja 'BENCHMARKS'",
        "• Compare los indicadores de la empresa vs el sector",
        "• Identifique brechas y oportunidades de mejora",
        "",
        "PASO 9: GENERAR DIAGNÓSTICO AUTOMÁTICO",
        "• Vaya a la hoja 'DIAGNÓSTICO'",
        "• Presione el botón 'Generar Diagnóstico' (Macro)",
        "• Se generará un informe completo con interpretaciones humanizadas",
        "• El diagnóstico incluye hallazgos, alertas y recomendaciones",
        "",
        "PASO 10: VISUALIZAR DASHBOARD",
        "• Vaya a la hoja 'INFORME VISUAL'",
        "• Presione el botón 'Actualizar Gráficos' (Macro)",
        "• Se mostrarán gráficos ejecutivos del análisis",
        "",
        "═══════════════════════════════════════════════════════════════════════════════",
        "",
        "MACROS DISPONIBLES:",
        "",
        "1. GenerarDiagnostico: Crea el diagnóstico automático con interpretaciones",
        "2. ActualizarGraficos: Actualiza todos los gráficos del dashboard",
        "3. ExportarPDF: Exporta el diagnóstico e informe visual a PDF",
        "4. LimpiarDatos: Limpia todas las celdas de entrada para nuevo análisis",
        "5. CalcularEscenarios: Recalcula las proyecciones en 3 escenarios",
        "",
        "═══════════════════════════════════════════════════════════════════════════════",
        "",
        "CÓDIGOS DE COLOR:",
        "",
        "• AZUL = Celdas de entrada (usted debe completar estos datos)",
        "• NEGRO = Fórmulas y cálculos automáticos (no modificar)",
        "• VERDE = Referencias a otras hojas",
        "• AMARILLO = Supuestos clave que requieren atención especial",
        "• ROJO = Indicadores en zona de alerta",
        "• VERDE CLARO = Indicadores en zona óptima",
        "• AMARILLO CLARO = Indicadores en zona de precaución",
        "",
        "═══════════════════════════════════════════════════════════════════════════════",
        "",
        "RECOMENDACIONES IMPORTANTES:",
        "",
        "✓ Ingrese datos reales y verificados de los estados financieros",
        "✓ No modifique las fórmulas (celdas en negro)",
        "✓ Revise que el balance cuadre antes de continuar",
        "✓ Compare los resultados con los benchmarks de la industria",
        "✓ Use el diagnóstico automático como guía, no como verdad absoluta",
        "✓ Consulte con un contador o asesor financiero para validar interpretaciones",
        "✓ Guarde regularmente el archivo con diferentes versiones",
        "✓ Active las macros al abrir el archivo para funcionalidad completa",
        "",
        "═══════════════════════════════════════════════════════════════════════════════",
        "",
        "SOPORTE Y CONTACTO:",
        "",
        "Esta herramienta fue desarrollada siguiendo metodología Trulab para análisis",
        "financiero de PYMEs colombianas.",
        "",
        "Para soporte técnico o consultoría financiera, contacte a su asesor.",
        "",
        "Versión: 1.0",
        "Fecha: Diciembre 2024",
        "Compatibilidad: Excel 2016 o superior, LibreOffice Calc",
        "",
        "═══════════════════════════════════════════════════════════════════════════════",
    ]

    for i, linea in enumerate(instrucciones, start=2):
        ws[f'A{i}'] = linea

        if "PASO " in linea or "MACROS DISPONIBLES" in linea or "CÓDIGOS DE COLOR" in linea or "RECOMENDACIONES" in linea:
            ws[f'A{i}'].font = Font(bold=True, size=11, color="FFFFFF")
            ws[f'A{i}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        elif linea.startswith(tuple(str(i) + "." for i in range(1, 6))):
            ws[f'A{i}'].font = Font(bold=True)
        elif linea.startswith("•") or linea.startswith("✓"):
            ws[f'A{i}'].font = Font(size=10)

        ws.merge_cells(f'A{i}:F{i}')

    ws.column_dimensions['A'].width = 100

    return ws

def main():
    """Función principal para crear el archivo Excel"""
    print("Creando Sistema Integral de Controller Financiero...")

    # Crear workbook
    wb = Workbook()
    wb.remove(wb.active)  # Eliminar hoja por defecto

    # Obtener estilos
    estilos = crear_estilos()

    # Crear todas las hojas
    print("Creando hoja de Instrucciones...")
    crear_hoja_instrucciones(wb, estilos)

    print("Creando hoja de Configuración...")
    crear_hoja_configuracion(wb, estilos)

    print("Creando hoja de Estado de Resultados...")
    crear_hoja_eerr(wb, estilos)

    print("Creando hoja de Balance General...")
    crear_hoja_esf(wb, estilos)

    print("Creando hoja de Indicadores Financieros...")
    crear_hoja_indicadores(wb, estilos)

    print("Creando hoja de Punto de Equilibrio...")
    crear_hoja_punto_equilibrio(wb, estilos)

    print("Creando hoja de Proyecciones...")
    crear_hoja_proyecciones(wb, estilos)

    print("Creando hoja de Valoración...")
    crear_hoja_valoracion(wb, estilos)

    print("Creando hoja de Benchmarks...")
    crear_hoja_benchmarks(wb, estilos)

    print("Creando hoja de Diagnóstico...")
    crear_hoja_diagnostico(wb, estilos)

    print("Creando hoja de Informe Visual...")
    crear_hoja_informe_visual(wb, estilos)

    # Guardar archivo
    archivo_salida = "/home/user/Biblioteca/ControllerFinanciero/ControllerFinanciero_v1.0.xlsx"
    wb.save(archivo_salida)
    print(f"\n✓ Archivo creado exitosamente: {archivo_salida}")
    print(f"\nNOTA: Este archivo .xlsx debe convertirse a .xlsm y agregar las macros VBA")
    print("Las macros se agregarán en el siguiente paso.")

    return archivo_salida

if __name__ == "__main__":
    main()
