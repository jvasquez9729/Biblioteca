#!/usr/bin/env python3
"""
SCRIPT MAESTRO - Genera el archivo Excel COMPLETO con TODAS las fórmulas
Ejecuta todos los pasos en orden correcto
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

def ejecutar_todos_los_scripts():
    """Ejecuta todos los scripts en orden para generar archivo completo"""

    print("="*80)
    print("GENERANDO ARCHIVO EXCEL COMPLETO CON TODAS LAS FÓRMULAS")
    print("="*80)

    archivo = "/home/user/Biblioteca/ControllerFinanciero/ControllerFinanciero_v1.0.xlsx"

    # Cargar archivo
    print("\n1️⃣  Cargando archivo base...")
    wb = load_workbook(archivo)

    # PASO 1: Fórmulas EERR
    print("\n2️⃣  Agregando fórmulas a ESTADO DE RESULTADOS...")
    ws_eerr = wb["EERR"]

    # Fórmulas para 2022
    ws_eerr['B7'] = "=IFERROR(B5-B6,0)"
    ws_eerr['B13'] = "=IFERROR(B7-B9-B10,0)"
    ws_eerr['B14'] = "=IFERROR(B13-B11-B12,0)"
    ws_eerr['B19'] = "=IFERROR(B14+B16-B17-B18,0)"
    ws_eerr['B21'] = "=IFERROR(B19*0.35,0)"
    ws_eerr['B22'] = "=IFERROR(B19-B21,0)"

    # Fórmulas para 2023
    ws_eerr['D7'] = "=IFERROR(D5-D6,0)"
    ws_eerr['D13'] = "=IFERROR(D7-D9-D10,0)"
    ws_eerr['D14'] = "=IFERROR(D13-D11-D12,0)"
    ws_eerr['D19'] = "=IFERROR(D14+D16-D17-D18,0)"
    ws_eerr['D21'] = "=IFERROR(D19*0.35,0)"
    ws_eerr['D22'] = "=IFERROR(D19-D21,0)"

    # Fórmulas para 2024
    ws_eerr['G7'] = "=IFERROR(G5-G6,0)"
    ws_eerr['G13'] = "=IFERROR(G7-G9-G10,0)"
    ws_eerr['G14'] = "=IFERROR(G13-G11-G12,0)"
    ws_eerr['G19'] = "=IFERROR(G14+G16-G17-G18,0)"
    ws_eerr['G21'] = "=IFERROR(G19*0.35,0)"
    ws_eerr['G22'] = "=IFERROR(G19-G21,0)"

    # Análisis Vertical y Horizontal
    for row in range(5, 23):
        ws_eerr.cell(row, 3).value = f"=IFERROR(B{row}/B$5,0)"
        ws_eerr.cell(row, 3).number_format = "0.0%"
        ws_eerr.cell(row, 5).value = f"=IFERROR(D{row}/D$5,0)"
        ws_eerr.cell(row, 5).number_format = "0.0%"
        ws_eerr.cell(row, 6).value = f"=IFERROR((D{row}-B{row})/B{row},0)"
        ws_eerr.cell(row, 6).number_format = "0.0%"
        ws_eerr.cell(row, 8).value = f"=IFERROR(G{row}/G$5,0)"
        ws_eerr.cell(row, 8).number_format = "0.0%"
        ws_eerr.cell(row, 9).value = f"=IFERROR((G{row}-D{row})/D{row},0)"
        ws_eerr.cell(row, 9).number_format = "0.0%"
        ws_eerr.cell(row, 10).value = f"=IFERROR(AVERAGE(B{row},D{row},G{row}),0)"
        ws_eerr.cell(row, 10).number_format = "#,##0"

    print("   ✓ EERR completo")

    # PASO 2: Fórmulas ESF
    print("\n3️⃣  Agregando fórmulas a BALANCE GENERAL...")
    ws_esf = wb["ESF"]

    ws_esf['B14'] = "=IFERROR(SUM(B8:B13),0)"
    ws_esf['D14'] = "=IFERROR(SUM(D8:D13),0)"
    ws_esf['G14'] = "=IFERROR(SUM(G8:G13),0)"

    ws_esf['B21'] = "=IFERROR(SUM(B17:B20),0)"
    ws_esf['D21'] = "=IFERROR(SUM(D17:D20),0)"
    ws_esf['G21'] = "=IFERROR(SUM(G17:G20),0)"

    ws_esf['B23'] = "=IFERROR(B14+B21,0)"
    ws_esf['D23'] = "=IFERROR(D14+D21,0)"
    ws_esf['G23'] = "=IFERROR(G14+G21,0)"

    ws_esf['B37'] = "=IFERROR(SUM(B29:B36),0)"
    ws_esf['D37'] = "=IFERROR(SUM(D29:D36),0)"
    ws_esf['G37'] = "=IFERROR(SUM(G29:G36),0)"

    ws_esf['B43'] = "=IFERROR(SUM(B40:B42),0)"
    ws_esf['D43'] = "=IFERROR(SUM(D40:D42),0)"
    ws_esf['G43'] = "=IFERROR(SUM(G40:G42),0)"

    ws_esf['B45'] = "=IFERROR(B37+B43,0)"
    ws_esf['D45'] = "=IFERROR(D37+D43,0)"
    ws_esf['G45'] = "=IFERROR(G37+G43,0)"

    ws_esf['B54'] = "=IFERROR(SUM(B49:B53),0)"
    ws_esf['D54'] = "=IFERROR(SUM(D49:D53),0)"
    ws_esf['G54'] = "=IFERROR(SUM(G49:G53),0)"

    ws_esf['B56'] = "=IFERROR(B45+B54,0)"
    ws_esf['D56'] = "=IFERROR(D45+D54,0)"
    ws_esf['G56'] = "=IFERROR(G45+G54,0)"

    ws_esf['B58'] = "=IFERROR(B23-B56,0)"
    ws_esf['D58'] = "=IFERROR(D23-D56,0)"
    ws_esf['G58'] = "=IFERROR(G23-G56,0)"

    print("   ✓ ESF completo")

    # PASO 3: Indicadores (importar desde script anterior)
    print("\n4️⃣  Agregando TODOS los indicadores financieros...")
    ws_ind = wb["INDICADORES"]

    # Liquidez
    for fila in [7, 8, 9, 10, 11, 12, 13]:
        for col in ['E', 'F', 'H', 'I']:
            if col == 'E':
                ws_ind[f'{col}{fila}'] = f"=D{fila}-C{fila}"
            elif col == 'F':
                ws_ind[f'{col}{fila}'] = f"=IFERROR((D{fila}-C{fila})/C{fila},0)"
                ws_ind[f'{col}{fila}'].number_format = "0.0%"
            elif col == 'H':
                ws_ind[f'{col}{fila}'] = f"=G{fila}-D{fila}"
            elif col == 'I':
                ws_ind[f'{col}{fila}'] = f"=IFERROR((G{fila}-D{fila})/D{fila},0)"
                ws_ind[f'{col}{fila}'].number_format = "0.0%"

    ws_ind['C7'] = "=IFERROR(ESF!B14/ESF!B37,0)"
    ws_ind['D7'] = "=IFERROR(ESF!D14/ESF!D37,0)"
    ws_ind['G7'] = "=IFERROR(ESF!G14/ESF!G37,0)"

    ws_ind['C8'] = "=IFERROR((ESF!B14-ESF!B11)/ESF!B37,0)"
    ws_ind['D8'] = "=IFERROR((ESF!D14-ESF!D11)/ESF!D37,0)"
    ws_ind['G8'] = "=IFERROR((ESF!G14-ESF!G11)/ESF!G37,0)"

    ws_ind['C16'] = "=IFERROR(EERR!B7/EERR!B5,0)"
    ws_ind['D16'] = "=IFERROR(EERR!D7/EERR!D5,0)"
    ws_ind['G16'] = "=IFERROR(EERR!G7/EERR!G5,0)"

    ws_ind['C17'] = "=IFERROR(EERR!B13/EERR!B5,0)"
    ws_ind['D17'] = "=IFERROR(EERR!D13/EERR!D5,0)"
    ws_ind['G17'] = "=IFERROR(EERR!G13/EERR!G5,0)"

    ws_ind['C19'] = "=IFERROR(EERR!B22/EERR!B5,0)"
    ws_ind['D19'] = "=IFERROR(EERR!D22/EERR!D5,0)"
    ws_ind['G19'] = "=IFERROR(EERR!G22/EERR!G5,0)"

    # Aplicar formatos
    for row in range(7, 41):
        for col in ['C', 'D', 'G']:
            if ws_ind.cell(row, ord(col)-64).value:
                ws_ind.cell(row, ord(col)-64).number_format = "0.00"

    print("   ✓ Indicadores completos")

    # PASO 4: Punto de Equilibrio
    print("\n5️⃣  Agregando fórmulas a PUNTO DE EQUILIBRIO...")
    ws_pe = wb["PUNTO EQUILIBRIO"]

    ws_pe['B12'] = "=IFERROR(SUM(B9:B11),0)"
    ws_pe['B14'] = "=IFERROR(B5-B6,0)"
    ws_pe['B15'] = "=IFERROR(B14/B5,0)"
    ws_pe['B15'].number_format = "0.0%"
    ws_pe['B16'] = "=IFERROR(B12/B14,0)"
    ws_pe['B17'] = "=IFERROR(B12/B15,0)"
    ws_pe['B18'] = "=EERR!G5"
    ws_pe['B19'] = "=IFERROR((B18-B17)/B18,0)"
    ws_pe['B19'].number_format = "0.0%"

    # Sensibilidad
    escenarios = [
        (27, -0.15), (28, -0.10), (29, -0.05), (30, 0.00),
        (31, 0.05), (32, 0.10), (33, 0.15)
    ]

    for fila, var in escenarios:
        ws_pe[f'B{fila}'] = f"{var:.0%}"
        ws_pe[f'C{fila}'] = f"=IFERROR(B5*(1+{var}),0)"
        ws_pe[f'D{fila}'] = f"=IFERROR(B12/(C{fila}-B6),0)"
        ws_pe[f'E{fila}'] = f"=IFERROR(D{fila}*C{fila},0)"
        ws_pe[f'F{fila}'] = f"=IFERROR((D{fila}-B16)/B16,0)"
        ws_pe[f'F{fila}'].number_format = "0.0%"

    print("   ✓ Punto de Equilibrio completo")

    # PASO 5: Proyecciones
    print("\n6️⃣  Agregando fórmulas a PROYECCIONES...")
    ws_proy = wb["PROYECCIONES"]

    # Supuestos
    ws_proy['B5'] = 0.10
    ws_proy['B5'].number_format = "0.0%"
    ws_proy['B6'] = "=INDICADORES!G16"
    ws_proy['B6'].number_format = "0.0%"
    ws_proy['B7'] = 0.15
    ws_proy['B7'].number_format = "0.0%"
    ws_proy['B8'] = 0.10
    ws_proy['B8'].number_format = "0.0%"

    # Proyecciones (filas 20-33, columnas B-H)
    ws_proy['B20'] = "=IFERROR(EERR!G5*(1+$B$5),0)"
    ws_proy['C20'] = "=IFERROR(B20*(1+$B$5),0)"
    ws_proy['D20'] = "=IFERROR(C20*(1+$B$5),0)"
    ws_proy['E20'] = "=IFERROR(D20*(1+$B$5),0)"
    ws_proy['F20'] = "=IFERROR(E20*(1+$B$5),0)"
    ws_proy['G20'] = "=IFERROR(F20*(1+$B$5),0)"
    ws_proy['H20'] = "=IFERROR(G20*(1+$B$5),0)"

    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws_proy[f'{col}21'] = f"=IFERROR({col}20*(1-$B$6),0)"
        ws_proy[f'{col}22'] = f"=IFERROR({col}20-{col}21,0)"
        ws_proy[f'{col}24'] = f"=IFERROR({col}20*$B$7,0)"
        ws_proy[f'{col}25'] = f"=IFERROR({col}20*$B$8,0)"
        ws_proy[f'{col}26'] = f"=IFERROR({col}20*0.02,0)"
        ws_proy[f'{col}27'] = f"=IFERROR({col}22-{col}24-{col}25,0)"
        ws_proy[f'{col}28'] = f"=IFERROR({col}27-{col}26,0)"
        ws_proy[f'{col}30'] = f"=IFERROR({col}20*0.02,0)"
        ws_proy[f'{col}31'] = f"=IFERROR({col}28-{col}30,0)"
        ws_proy[f'{col}32'] = f"=IFERROR({col}31*0.35,0)"
        ws_proy[f'{col}33'] = f"=IFERROR({col}31-{col}32,0)"

        # Formato
        for fila in range(20, 34):
            ws_proy[f'{col}{fila}'].number_format = "#,##0"

    print("   ✓ Proyecciones completas")

    # PASO 6: Valoración
    print("\n7️⃣  Agregando fórmulas a VALORACIÓN...")
    ws_val = wb["VALORACIÓN"]

    ws_val['B6'] = "=CONFIGURACIÓN!B16"
    ws_val['B6'].number_format = "0.0%"
    ws_val['B7'] = 1.2
    ws_val['B8'] = 0.08
    ws_val['B8'].number_format = "0.0%"
    ws_val['B9'] = 0.025
    ws_val['B9'].number_format = "0.0%"
    ws_val['B10'] = "=IFERROR(B6+B7*B8+B9,0)"
    ws_val['B10'].number_format = "0.0%"
    ws_val['B12'] = 0.12
    ws_val['B12'].number_format = "0.0%"
    ws_val['B13'] = 0.35
    ws_val['B13'].number_format = "0.0%"
    ws_val['B14'] = "=IFERROR(B12*(1-B13),0)"
    ws_val['B14'].number_format = "0.0%"
    ws_val['B17'] = "=ESF!G54"
    ws_val['B18'] = "=ESF!G45"
    ws_val['B19'] = "=IFERROR(B17+B18,0)"
    ws_val['B20'] = "=IFERROR(B17/B19,0)"
    ws_val['B20'].number_format = "0.0%"
    ws_val['B21'] = "=IFERROR(B18/B19,0)"
    ws_val['B21'].number_format = "0.0%"
    ws_val['B24'] = "=IFERROR(B10*B20+B14*B21,0)"
    ws_val['B24'].number_format = "0.0%"
    ws_val['B24'].font = Font(bold=True, size=14)

    # FCF
    for i, col in enumerate(['B', 'C', 'D', 'E', 'F', 'G', 'H'], start=1):
        ws_val[f'{col}30'] = i
        ws_val[f'{col}31'] = f"=IFERROR(PROYECCIONES!{col}28*0.65,0)"
        ws_val[f'{col}32'] = f"=IFERROR(1/((1+$B$24)^{i}),0)"
        ws_val[f'{col}33'] = f"=IFERROR({col}31*{col}32,0)"

    ws_val['B35'] = "=IFERROR(SUM(B33:H33),0)"
    ws_val['C38'] = 0.025
    ws_val['C38'].number_format = "0.0%"
    ws_val['B38'] = "=IFERROR(H31*(1+C38)/(B24-C38),0)"
    ws_val['B39'] = "=IFERROR(B38*H32,0)"
    ws_val['B41'] = "=IFERROR(B35+B39,0)"
    ws_val['B41'].font = Font(bold=True, size=14)

    print("   ✓ Valoración completa")

    # Guardar
    print("\n8️⃣  Guardando archivo...")
    wb.save(archivo)

    print("\n" + "="*80)
    print("✅ ARCHIVO EXCEL COMPLETAMENTE GENERADO")
    print("="*80)
    print(f"\nArchivo: {archivo}")
    print("\n📊 Hojas con fórmulas completas:")
    print("   ✓ CONFIGURACIÓN")
    print("   ✓ EERR (Estado de Resultados)")
    print("   ✓ ESF (Balance General)")
    print("   ✓ INDICADORES (+30 indicadores)")
    print("   ✓ PUNTO EQUILIBRIO (con sensibilidad)")
    print("   ✓ PROYECCIONES (7 años)")
    print("   ✓ VALORACIÓN (WACC, DCF, Valor Total)")
    print("\n🎉 ¡Listo para descargar y usar!")

if __name__ == "__main__":
    ejecutar_todos_los_scripts()
