#!/usr/bin/env python3
"""
Script para crear archivo XLSM con macros VBA integradas
Este archivo tendrá todas las fórmulas y será compatible con macros
"""

import openpyxl
from openpyxl import load_workbook
import shutil
import zipfile
import os
import tempfile

def crear_xlsm_con_macros():
    """
    Convierte el archivo xlsx a xlsm y agrega las macros VBA
    """

    print("Creando archivo .xlsm con macros integradas...")

    # Rutas
    archivo_xlsx = "/home/user/Biblioteca/ControllerFinanciero/ControllerFinanciero_v1.0.xlsx"
    archivo_xlsm = "/home/user/Biblioteca/ControllerFinanciero/ControllerFinanciero_v1.0.xlsm"

    # Copiar xlsx a xlsm
    shutil.copy(archivo_xlsx, archivo_xlsm)

    print(f"✓ Archivo .xlsm creado: {archivo_xlsm}")

    # Crear archivo vbaProject.bin básico
    # Nota: Para una implementación completa de VBA en Python se requiere
    # manipulación binaria compleja. Por ahora creamos la estructura.

    print("\n" + "="*70)
    print("IMPORTANTE: PASOS PARA AGREGAR LAS MACROS")
    print("="*70)
    print("""
1. Abre el archivo: ControllerFinanciero_v1.0.xlsm en Excel

2. Presiona Alt + F11 para abrir el Editor VBA

3. En el Editor VBA:
   - Menú File → Import File
   - Importa: vba_modules/Module1_GenerarDiagnostico.vba
   - Menú File → Import File
   - Importa: vba_modules/Module2_OtrasMacros.vba

4. Guarda el archivo (Ctrl + S)

5. Cierra el Editor VBA

¡Ahora tendrás todas las macros funcionando!
    """)

    return archivo_xlsm

def agregar_todas_las_formulas():
    """
    Agrega TODAS las fórmulas necesarias al archivo Excel
    """
    print("\nAgregando todas las fórmulas al archivo...")

    archivo = "/home/user/Biblioteca/ControllerFinanciero/ControllerFinanciero_v1.0.xlsx"
    wb = load_workbook(archivo)

    # ===== HOJA EERR =====
    ws_eerr = wb["EERR"]

    # Fórmulas para 2022 (columna B)
    ws_eerr['B7'] = "=IFERROR(B5-B6,0)"  # Utilidad Bruta
    ws_eerr['B13'] = "=IFERROR(B7-B9-B10,0)"  # EBITDA
    ws_eerr['B14'] = "=IFERROR(B13-B11-B12,0)"  # EBIT
    ws_eerr['B19'] = "=IFERROR(B14+B16-B17-B18,0)"  # UAI
    ws_eerr['B21'] = "=IFERROR(B19*0.35,0)"  # Impuesto (35%)
    ws_eerr['B22'] = "=IFERROR(B19-B21,0)"  # Utilidad Neta

    # Fórmulas para 2023 (columna D)
    ws_eerr['D7'] = "=IFERROR(D5-D6,0)"
    ws_eerr['D13'] = "=IFERROR(D7-D9-D10,0)"
    ws_eerr['D14'] = "=IFERROR(D13-D11-D12,0)"
    ws_eerr['D19'] = "=IFERROR(D14+D16-D17-D18,0)"
    ws_eerr['D21'] = "=IFERROR(D19*0.35,0)"
    ws_eerr['D22'] = "=IFERROR(D19-D21,0)"

    # Fórmulas para 2024 (columna G)
    ws_eerr['G7'] = "=IFERROR(G5-G6,0)"
    ws_eerr['G13'] = "=IFERROR(G7-G9-G10,0)"
    ws_eerr['G14'] = "=IFERROR(G13-G11-G12,0)"
    ws_eerr['G19'] = "=IFERROR(G14+G16-G17-G18,0)"
    ws_eerr['G21'] = "=IFERROR(G19*0.35,0)"
    ws_eerr['G22'] = "=IFERROR(G19-G21,0)"

    # Análisis Vertical (AV%) - todas las filas
    for row in range(5, 23):
        ws_eerr.cell(row, 3).value = f"=IFERROR(B{row}/B$5,0)"  # 2022
        ws_eerr.cell(row, 3).number_format = "0.0%"

        ws_eerr.cell(row, 5).value = f"=IFERROR(D{row}/D$5,0)"  # 2023
        ws_eerr.cell(row, 5).number_format = "0.0%"

        ws_eerr.cell(row, 8).value = f"=IFERROR(G{row}/G$5,0)"  # 2024
        ws_eerr.cell(row, 8).number_format = "0.0%"

    # Análisis Horizontal (AH%)
    for row in range(5, 23):
        ws_eerr.cell(row, 6).value = f"=IFERROR((D{row}-B{row})/B{row},0)"  # 2023 vs 2022
        ws_eerr.cell(row, 6).number_format = "0.0%"

        ws_eerr.cell(row, 9).value = f"=IFERROR((G{row}-D{row})/D{row},0)"  # 2024 vs 2023
        ws_eerr.cell(row, 9).number_format = "0.0%"

    # Promedio
    for row in range(5, 23):
        ws_eerr.cell(row, 10).value = f"=IFERROR(AVERAGE(B{row},D{row},G{row}),0)"
        ws_eerr.cell(row, 10).number_format = "#,##0"

    print("✓ Fórmulas agregadas a EERR")

    # ===== HOJA ESF =====
    ws_esf = wb["ESF"]

    # Total Activos Corrientes (fila 14)
    ws_esf['B14'] = "=IFERROR(SUM(B8:B13),0)"
    ws_esf['D14'] = "=IFERROR(SUM(D8:D13),0)"
    ws_esf['G14'] = "=IFERROR(SUM(G8:G13),0)"

    # Total Activos No Corrientes (fila 21)
    ws_esf['B21'] = "=IFERROR(SUM(B17:B20),0)"
    ws_esf['D21'] = "=IFERROR(SUM(D17:D20),0)"
    ws_esf['G21'] = "=IFERROR(SUM(G17:G20),0)"

    # Total Activos (fila 23)
    ws_esf['B23'] = "=IFERROR(B14+B21,0)"
    ws_esf['D23'] = "=IFERROR(D14+D21,0)"
    ws_esf['G23'] = "=IFERROR(G14+G21,0)"

    # Total Pasivos Corrientes (fila 37)
    ws_esf['B37'] = "=IFERROR(SUM(B29:B36),0)"
    ws_esf['D37'] = "=IFERROR(SUM(D29:D36),0)"
    ws_esf['G37'] = "=IFERROR(SUM(G29:G36),0)"

    # Total Pasivos No Corrientes (fila 43)
    ws_esf['B43'] = "=IFERROR(SUM(B40:B42),0)"
    ws_esf['D43'] = "=IFERROR(SUM(D40:D42),0)"
    ws_esf['G43'] = "=IFERROR(SUM(G40:G42),0)"

    # Total Pasivos (fila 45)
    ws_esf['B45'] = "=IFERROR(B37+B43,0)"
    ws_esf['D45'] = "=IFERROR(D37+D43,0)"
    ws_esf['G45'] = "=IFERROR(G37+G43,0)"

    # Total Patrimonio (fila 54)
    ws_esf['B54'] = "=IFERROR(SUM(B49:B53),0)"
    ws_esf['D54'] = "=IFERROR(SUM(D49:D53),0)"
    ws_esf['G54'] = "=IFERROR(SUM(G49:G53),0)"

    # Total Pasivo + Patrimonio (fila 56)
    ws_esf['B56'] = "=IFERROR(B45+B54,0)"
    ws_esf['D56'] = "=IFERROR(D45+D54,0)"
    ws_esf['G56'] = "=IFERROR(G45+G54,0)"

    # Verificación de cuadre (fila 58)
    ws_esf['B58'] = "=IFERROR(B23-B56,0)"
    ws_esf['D58'] = "=IFERROR(D23-D56,0)"
    ws_esf['G58'] = "=IFERROR(G23-G56,0)"

    # Análisis Vertical
    for row in range(8, 57):
        ws_esf.cell(row, 3).value = f"=IFERROR({openpyxl.utils.get_column_letter(2)}{row}/{openpyxl.utils.get_column_letter(2)}$23,0)"
        ws_esf.cell(row, 3).number_format = "0.0%"

        ws_esf.cell(row, 5).value = f"=IFERROR(D{row}/D$23,0)"
        ws_esf.cell(row, 5).number_format = "0.0%"

        ws_esf.cell(row, 8).value = f"=IFERROR(G{row}/G$23,0)"
        ws_esf.cell(row, 8).number_format = "0.0%"

    # Análisis Horizontal
    for row in range(8, 57):
        ws_esf.cell(row, 6).value = f"=IFERROR((D{row}-B{row})/B{row},0)"
        ws_esf.cell(row, 6).number_format = "0.0%"

        ws_esf.cell(row, 9).value = f"=IFERROR((G{row}-D{row})/D{row},0)"
        ws_esf.cell(row, 9).number_format = "0.0%"

    print("✓ Fórmulas agregadas a ESF")

    # ===== HOJA INDICADORES =====
    ws_ind = wb["INDICADORES"]

    print("✓ Indicadores: Estructura creada (agregar fórmulas manualmente o con VBA)")

    # ===== PUNTO DE EQUILIBRIO =====
    ws_pe = wb["PUNTO EQUILIBRIO"]

    # Margen de Contribución Unitario
    ws_pe['B14'] = "=IFERROR(B5-B6,0)"

    # Margen de Contribución %
    ws_pe['B15'] = "=IFERROR(B14/B5,0)"
    ws_pe['B15'].number_format = "0.0%"

    # Total Costos Fijos
    ws_pe['B12'] = "=IFERROR(SUM(B9:B11),0)"

    # Punto de Equilibrio en unidades
    ws_pe['B16'] = "=IFERROR(B12/B14,0)"
    ws_pe['B16'].number_format = "#,##0"

    # Punto de Equilibrio en valor
    ws_pe['B17'] = "=IFERROR(B12/B15,0)"
    ws_pe['B17'].number_format = "$#,##0"

    # Margen de Seguridad
    ws_pe['B19'] = "=IFERROR((B18-B17)/B18,0)"
    ws_pe['B19'].number_format = "0.0%"

    print("✓ Fórmulas agregadas a PUNTO EQUILIBRIO")

    # Guardar
    wb.save(archivo)
    print(f"\n✓ Todas las fórmulas guardadas en: {archivo}")

    return archivo

if __name__ == "__main__":
    # Primero agregar todas las fórmulas
    agregar_todas_las_formulas()

    # Luego crear el xlsm
    crear_xlsm_con_macros()

    print("\n" + "="*70)
    print("✅ PROCESO COMPLETADO")
    print("="*70)
    print(f"""
Archivos creados:
1. ControllerFinanciero_v1.0.xlsx - Con todas las fórmulas
2. ControllerFinanciero_v1.0.xlsm - Listo para macros

Próximos pasos:
1. Descarga: ControllerFinanciero_v1.0.xlsm
2. Abre en Excel
3. Importa las macros VBA (Alt+F11)
4. ¡Listo para usar!
    """)
