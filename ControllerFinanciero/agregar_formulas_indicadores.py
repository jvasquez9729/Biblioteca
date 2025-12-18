#!/usr/bin/env python3
"""
Script para agregar TODAS las fórmulas de indicadores financieros
"""

import openpyxl
from openpyxl import load_workbook

def agregar_formulas_indicadores():
    """Agrega todas las fórmulas a la hoja de indicadores"""

    print("Agregando fórmulas de indicadores financieros...")

    archivo = "/home/user/Biblioteca/ControllerFinanciero/ControllerFinanciero_v1.0.xlsx"
    wb = load_workbook(archivo)
    ws = wb["INDICADORES"]

    # A. INDICADORES DE LIQUIDEZ

    # Fila 7: Razón Corriente
    ws['C7'] = "=IFERROR(ESF!B14/ESF!B37,0)"
    ws['D7'] = "=IFERROR(ESF!D14/ESF!D37,0)"
    ws['G7'] = "=IFERROR(ESF!G14/ESF!G37,0)"
    ws['E7'] = "=D7-C7"
    ws['F7'] = "=IFERROR((D7-C7)/C7,0)"
    ws['H7'] = "=G7-D7"
    ws['I7'] = "=IFERROR((G7-D7)/D7,0)"
    ws['J7'] = '=IF(G7>=1.5,"🟢 Óptimo",IF(G7>=1,"🟡 Aceptable","🔴 Crítico"))'

    # Fila 8: Prueba Ácida
    ws['C8'] = "=IFERROR((ESF!B14-ESF!B11)/ESF!B37,0)"
    ws['D8'] = "=IFERROR((ESF!D14-ESF!D11)/ESF!D37,0)"
    ws['G8'] = "=IFERROR((ESF!G14-ESF!G11)/ESF!G37,0)"
    ws['E8'] = "=D8-C8"
    ws['F8'] = "=IFERROR((D8-C8)/C8,0)"
    ws['H8'] = "=G8-D8"
    ws['I8'] = "=IFERROR((G8-D8)/D8,0)"
    ws['J8'] = '=IF(G8>=1,"🟢 Óptimo",IF(G8>=0.8,"🟡 Aceptable","🔴 Crítico"))'

    # Fila 9: Capital de Trabajo (KT)
    ws['C9'] = "=IFERROR(ESF!B14-ESF!B37,0)"
    ws['D9'] = "=IFERROR(ESF!D14-ESF!D37,0)"
    ws['G9'] = "=IFERROR(ESF!G14-ESF!G37,0)"
    ws['E9'] = "=D9-C9"
    ws['F9'] = "=IFERROR((D9-C9)/C9,0)"
    ws['H9'] = "=G9-D9"
    ws['I9'] = "=IFERROR((G9-D9)/D9,0)"
    ws['J9'] = '=IF(G9>0,"🟢 Positivo","🔴 Negativo")'

    # Fila 10: KTNO
    ws['C10'] = "=IFERROR(ESF!B9+ESF!B11-ESF!B30,0)"
    ws['D10'] = "=IFERROR(ESF!D9+ESF!D11-ESF!D30,0)"
    ws['G10'] = "=IFERROR(ESF!G9+ESF!G11-ESF!G30,0)"
    ws['E10'] = "=D10-C10"
    ws['F10'] = "=IFERROR((D10-C10)/C10,0)"
    ws['H10'] = "=G10-D10"
    ws['I10'] = "=IFERROR((G10-D10)/D10,0)"

    # Fila 11: PKT
    ws['C11'] = "=IFERROR(C9/EERR!B5,0)"
    ws['D11'] = "=IFERROR(D9/EERR!D5,0)"
    ws['G11'] = "=IFERROR(G9/EERR!G5,0)"
    ws['E11'] = "=D11-C11"
    ws['F11'] = "=IFERROR((D11-C11)/C11,0)"
    ws['H11'] = "=G11-D11"
    ws['I11'] = "=IFERROR((G11-D11)/D11,0)"

    # Fila 12: PKTNO
    ws['C12'] = "=IFERROR(C10/EERR!B5,0)"
    ws['D12'] = "=IFERROR(D10/EERR!D5,0)"
    ws['G12'] = "=IFERROR(G10/EERR!G5,0)"
    ws['E12'] = "=D12-C12"
    ws['F12'] = "=IFERROR((D12-C12)/C12,0)"
    ws['H12'] = "=G12-D12"
    ws['I12'] = "=IFERROR((G12-D12)/D12,0)"

    # Fila 13: Solvencia
    ws['C13'] = "=IFERROR(ESF!B23/ESF!B45,0)"
    ws['D13'] = "=IFERROR(ESF!D23/ESF!D45,0)"
    ws['G13'] = "=IFERROR(ESF!G23/ESF!G45,0)"
    ws['E13'] = "=D13-C13"
    ws['F13'] = "=IFERROR((D13-C13)/C13,0)"
    ws['H13'] = "=G13-D13"
    ws['I13'] = "=IFERROR((G13-D13)/D13,0)"
    ws['J13'] = '=IF(G13>=2,"🟢 Óptimo",IF(G13>=1,"🟡 Aceptable","🔴 Crítico"))'

    print("✓ Indicadores de Liquidez")

    # B. INDICADORES DE RENTABILIDAD

    # Fila 16: Margen Bruto
    ws['C16'] = "=IFERROR(EERR!B7/EERR!B5,0)"
    ws['D16'] = "=IFERROR(EERR!D7/EERR!D5,0)"
    ws['G16'] = "=IFERROR(EERR!G7/EERR!G5,0)"
    ws['E16'] = "=D16-C16"
    ws['F16'] = "=IFERROR((D16-C16)/C16,0)"
    ws['H16'] = "=G16-D16"
    ws['I16'] = "=IFERROR((G16-D16)/D16,0)"
    ws['J16'] = '=IF(G16>=0.3,"🟢 Óptimo",IF(G16>=0.2,"🟡 Aceptable","🔴 Bajo"))'

    # Fila 17: Margen EBITDA
    ws['C17'] = "=IFERROR(EERR!B13/EERR!B5,0)"
    ws['D17'] = "=IFERROR(EERR!D13/EERR!D5,0)"
    ws['G17'] = "=IFERROR(EERR!G13/EERR!G5,0)"
    ws['E17'] = "=D17-C17"
    ws['F17'] = "=IFERROR((D17-C17)/C17,0)"
    ws['H17'] = "=G17-D17"
    ws['I17'] = "=IFERROR((G17-D17)/D17,0)"
    ws['J17'] = '=IF(G17>=0.15,"🟢 Óptimo",IF(G17>=0.1,"🟡 Aceptable","🔴 Bajo"))'

    # Fila 18: Margen Operativo
    ws['C18'] = "=IFERROR(EERR!B14/EERR!B5,0)"
    ws['D18'] = "=IFERROR(EERR!D14/EERR!D5,0)"
    ws['G18'] = "=IFERROR(EERR!G14/EERR!G5,0)"
    ws['E18'] = "=D18-C18"
    ws['F18'] = "=IFERROR((D18-C18)/C18,0)"
    ws['H18'] = "=G18-D18"
    ws['I18'] = "=IFERROR((G18-D18)/D18,0)"
    ws['J18'] = '=IF(G18>=0.1,"🟢 Óptimo",IF(G18>=0.05,"🟡 Aceptable","🔴 Bajo"))'

    # Fila 19: Margen Neto
    ws['C19'] = "=IFERROR(EERR!B22/EERR!B5,0)"
    ws['D19'] = "=IFERROR(EERR!D22/EERR!D5,0)"
    ws['G19'] = "=IFERROR(EERR!G22/EERR!G5,0)"
    ws['E19'] = "=D19-C19"
    ws['F19'] = "=IFERROR((D19-C19)/C19,0)"
    ws['H19'] = "=G19-D19"
    ws['I19'] = "=IFERROR((G19-D19)/D19,0)"
    ws['J19'] = '=IF(G19>=0.1,"🟢 Óptimo",IF(G19>=0.05,"🟡 Aceptable","🔴 Bajo"))'

    # Fila 20: ROA
    ws['C20'] = "=IFERROR(EERR!B22/ESF!B23,0)"
    ws['D20'] = "=IFERROR(EERR!D22/ESF!D23,0)"
    ws['G20'] = "=IFERROR(EERR!G22/ESF!G23,0)"
    ws['E20'] = "=D20-C20"
    ws['F20'] = "=IFERROR((D20-C20)/C20,0)"
    ws['H20'] = "=G20-D20"
    ws['I20'] = "=IFERROR((G20-D20)/D20,0)"
    ws['J20'] = '=IF(G20>=0.1,"🟢 Óptimo",IF(G20>=0.05,"🟡 Aceptable","🔴 Bajo"))'

    # Fila 21: ROE
    ws['C21'] = "=IFERROR(EERR!B22/ESF!B54,0)"
    ws['D21'] = "=IFERROR(EERR!D22/ESF!D54,0)"
    ws['G21'] = "=IFERROR(EERR!G22/ESF!G54,0)"
    ws['E21'] = "=D21-C21"
    ws['F21'] = "=IFERROR((D21-C21)/C21,0)"
    ws['H21'] = "=G21-D21"
    ws['I21'] = "=IFERROR((G21-D21)/D21,0)"
    ws['J21'] = '=IF(G21>=0.15,"🟢 Óptimo",IF(G21>=0.1,"🟡 Aceptable","🔴 Bajo"))'

    # Fila 22: Tasa Efectiva de Tributación
    ws['C22'] = "=IFERROR(EERR!B21/EERR!B19,0)"
    ws['D22'] = "=IFERROR(EERR!D21/EERR!D19,0)"
    ws['G22'] = "=IFERROR(EERR!G21/EERR!G19,0)"
    ws['E22'] = "=D22-C22"
    ws['F22'] = "=IFERROR((D22-C22)/C22,0)"
    ws['H22'] = "=G22-D22"
    ws['I22'] = "=IFERROR((G22-D22)/D22,0)"

    print("✓ Indicadores de Rentabilidad")

    # C. INDICADORES DE ENDEUDAMIENTO

    # Fila 25: Nivel de Endeudamiento
    ws['C25'] = "=IFERROR(ESF!B45/ESF!B23,0)"
    ws['D25'] = "=IFERROR(ESF!D45/ESF!D23,0)"
    ws['G25'] = "=IFERROR(ESF!G45/ESF!G23,0)"
    ws['E25'] = "=D25-C25"
    ws['F25'] = "=IFERROR((D25-C25)/C25,0)"
    ws['H25'] = "=G25-D25"
    ws['I25'] = "=IFERROR((G25-D25)/D25,0)"
    ws['J25'] = '=IF(G25<=0.5,"🟢 Óptimo",IF(G25<=0.7,"🟡 Moderado","🔴 Alto"))'

    # Fila 26: Concentración Deuda CP
    ws['C26'] = "=IFERROR(ESF!B37/ESF!B45,0)"
    ws['D26'] = "=IFERROR(ESF!D37/ESF!D45,0)"
    ws['G26'] = "=IFERROR(ESF!G37/ESF!G45,0)"
    ws['E26'] = "=D26-C26"
    ws['F26'] = "=IFERROR((D26-C26)/C26,0)"
    ws['H26'] = "=G26-D26"
    ws['I26'] = "=IFERROR((G26-D26)/D26,0)"

    # Fila 27: Deuda Financiera / Activos
    ws['C27'] = "=IFERROR((ESF!B29+ESF!B40)/ESF!B23,0)"
    ws['D27'] = "=IFERROR((ESF!D29+ESF!D40)/ESF!D23,0)"
    ws['G27'] = "=IFERROR((ESF!G29+ESF!G40)/ESF!G23,0)"
    ws['E27'] = "=D27-C27"
    ws['F27'] = "=IFERROR((D27-C27)/C27,0)"
    ws['H27'] = "=G27-D27"
    ws['I27'] = "=IFERROR((G27-D27)/D27,0)"

    # Fila 28: Deuda / Patrimonio
    ws['C28'] = "=IFERROR(ESF!B45/ESF!B54,0)"
    ws['D28'] = "=IFERROR(ESF!D45/ESF!D54,0)"
    ws['G28'] = "=IFERROR(ESF!G45/ESF!G54,0)"
    ws['E28'] = "=D28-C28"
    ws['F28'] = "=IFERROR((D28-C28)/C28,0)"
    ws['H28'] = "=G28-D28"
    ws['I28'] = "=IFERROR((G28-D28)/D28,0)"

    # Fila 29: Cobertura de Intereses
    ws['C29'] = "=IFERROR(EERR!B14/EERR!B18,0)"
    ws['D29'] = "=IFERROR(EERR!D14/EERR!D18,0)"
    ws['G29'] = "=IFERROR(EERR!G14/EERR!G18,0)"
    ws['E29'] = "=D29-C29"
    ws['F29'] = "=IFERROR((D29-C29)/C29,0)"
    ws['H29'] = "=G29-D29"
    ws['I29'] = "=IFERROR((G29-D29)/D29,0)"
    ws['J29'] = '=IF(G29>=5,"🟢 Óptimo",IF(G29>=2,"🟡 Aceptable","🔴 Crítico"))'

    print("✓ Indicadores de Endeudamiento")

    # D. INDICADORES DE ACTIVIDAD/ROTACIÓN

    # Fila 32: Rotación de Inventarios
    ws['C32'] = "=IFERROR(EERR!B6/ESF!B11,0)"
    ws['D32'] = "=IFERROR(EERR!D6/ESF!D11,0)"
    ws['G32'] = "=IFERROR(EERR!G6/ESF!G11,0)"
    ws['E32'] = "=D32-C32"
    ws['F32'] = "=IFERROR((D32-C32)/C32,0)"
    ws['H32'] = "=G32-D32"
    ws['I32'] = "=IFERROR((G32-D32)/D32,0)"

    # Fila 33: Días de Inventario
    ws['C33'] = "=IFERROR(365/C32,0)"
    ws['D33'] = "=IFERROR(365/D32,0)"
    ws['G33'] = "=IFERROR(365/G32,0)"
    ws['E33'] = "=D33-C33"
    ws['F33'] = "=IFERROR((D33-C33)/C33,0)"
    ws['H33'] = "=G33-D33"
    ws['I33'] = "=IFERROR((G33-D33)/D33,0)"

    # Fila 34: Rotación de Cartera
    ws['C34'] = "=IFERROR(EERR!B5/ESF!B9,0)"
    ws['D34'] = "=IFERROR(EERR!D5/ESF!D9,0)"
    ws['G34'] = "=IFERROR(EERR!G5/ESF!G9,0)"
    ws['E34'] = "=D34-C34"
    ws['F34'] = "=IFERROR((D34-C34)/C34,0)"
    ws['H34'] = "=G34-D34"
    ws['I34'] = "=IFERROR((G34-D34)/D34,0)"

    # Fila 35: Días de Cartera
    ws['C35'] = "=IFERROR(365/C34,0)"
    ws['D35'] = "=IFERROR(365/D34,0)"
    ws['G35'] = "=IFERROR(365/G34,0)"
    ws['E35'] = "=D35-C35"
    ws['F35'] = "=IFERROR((D35-C35)/C35,0)"
    ws['H35'] = "=G35-D35"
    ws['I35'] = "=IFERROR((G35-D35)/D35,0)"

    # Fila 36: Rotación de Proveedores (ajustada)
    ws['C36'] = "=IFERROR(EERR!B6/ESF!B30,0)"
    ws['D36'] = "=IFERROR(EERR!D6/ESF!D30,0)"
    ws['G36'] = "=IFERROR(EERR!G6/ESF!G30,0)"
    ws['E36'] = "=D36-C36"
    ws['F36'] = "=IFERROR((D36-C36)/C36,0)"
    ws['H36'] = "=G36-D36"
    ws['I36'] = "=IFERROR((G36-D36)/D36,0)"

    # Fila 37: Días de Proveedores
    ws['C37'] = "=IFERROR(365/C36,0)"
    ws['D37'] = "=IFERROR(365/D36,0)"
    ws['G37'] = "=IFERROR(365/G36,0)"
    ws['E37'] = "=D37-C37"
    ws['F37'] = "=IFERROR((D37-C37)/C37,0)"
    ws['H37'] = "=G37-D37"
    ws['I37'] = "=IFERROR((G37-D37)/D37,0)"

    # Fila 38: Ciclo Operativo
    ws['C38'] = "=IFERROR(C33+C35,0)"
    ws['D38'] = "=IFERROR(D33+D35,0)"
    ws['G38'] = "=IFERROR(G33+G35,0)"
    ws['E38'] = "=D38-C38"
    ws['F38'] = "=IFERROR((D38-C38)/C38,0)"
    ws['H38'] = "=G38-D38"
    ws['I38'] = "=IFERROR((G38-D38)/D38,0)"

    # Fila 39: Ciclo de Conversión de Efectivo
    ws['C39'] = "=IFERROR(C38-C37,0)"
    ws['D39'] = "=IFERROR(D38-D37,0)"
    ws['G39'] = "=IFERROR(G38-G37,0)"
    ws['E39'] = "=D39-C39"
    ws['F39'] = "=IFERROR((D39-C39)/C39,0)"
    ws['H39'] = "=G39-D39"
    ws['I39'] = "=IFERROR((G39-D39)/D39,0)"
    ws['J39'] = '=IF(G39<=30,"🟢 Óptimo",IF(G39<=60,"🟡 Aceptable","🔴 Alto"))'

    # Fila 40: GAP Financiero
    ws['C40'] = "=C39"
    ws['D40'] = "=D39"
    ws['G40'] = "=G39"
    ws['E40'] = "=D40-C40"
    ws['F40'] = "=IFERROR((D40-C40)/C40,0)"
    ws['H40'] = "=G40-D40"
    ws['I40'] = "=IFERROR((G40-D40)/D40,0)"

    print("✓ Indicadores de Actividad/Rotación")

    # Aplicar formatos de número
    for row in range(7, 41):
        for col in ['C', 'D', 'E', 'G', 'H']:
            cell = ws[f'{col}{row}']
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                # Formato decimal para la mayoría
                cell.number_format = "0.00"

        # Formato de porcentaje para variaciones
        for col in ['F', 'I']:
            cell = ws[f'{col}{row}']
            if cell.value:
                cell.number_format = "0.0%"

    # Formato especial de porcentaje para márgenes y ratios
    for row in [16, 17, 18, 19, 20, 21, 22, 25, 26, 27]:
        for col in ['C', 'D', 'G']:
            ws[f'{col}{row}'].number_format = "0.0%"

    # Formato de días (sin decimales)
    for row in [33, 35, 37, 38, 39, 40]:
        for col in ['C', 'D', 'G']:
            ws[f'{col}{row}'].number_format = "#,##0"

    print("✓ Formatos aplicados")

    # Guardar
    wb.save(archivo)
    print(f"\n✅ Todas las fórmulas de indicadores agregadas a: {archivo}")

if __name__ == "__main__":
    agregar_formulas_indicadores()
