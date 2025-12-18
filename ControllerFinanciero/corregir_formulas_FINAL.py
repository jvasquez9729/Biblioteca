#!/usr/bin/env python3
"""
Script para CORREGIR las fórmulas en las filas EXACTAS según la estructura real
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

archivo = "/home/user/Biblioteca/ControllerFinanciero/ControllerFinanciero_v1.0.xlsx"

print("="*80)
print("CORRIGIENDO FÓRMULAS EN FILAS EXACTAS")
print("="*80)

wb = load_workbook(archivo)

# ============================================================================
# PUNTO DE EQUILIBRIO - Las fórmulas ya están, solo falta sensibilidad
# ============================================================================
print("\n📊 Corrigiendo PUNTO DE EQUILIBRIO...")
ws_pe = wb["PUNTO EQUILIBRIO"]

# Las fórmulas principales ya están en filas 12, 14-19
# Solo necesito agregar las columnas C, D, E, F para el análisis de sensibilidad

# Encabezados de sensibilidad (fila 23)
ws_pe['C23'] = "Nuevo Precic PE Unidades"
ws_pe['D23'] = "PE Unidades"
ws_pe['E23'] = "PE Valor ($)"
ws_pe['F23'] = "Impacto"

# Escenarios (filas 24-30)
escenarios_sens = [
    (24, -0.15, "Precio -15%"),
    (25, -0.10, "Precio -10%"),
    (26, -0.05, "Precio -5%"),
    (27, 0.00, "Precio Base"),
    (28, 0.05, "Precio +5%"),
    (29, 0.10, "Precio +10%"),
    (30, 0.15, "Precio +15%"),
]

for fila, variacion, etiqueta in escenarios_sens:
    # Columna B ya tiene el porcentaje, solo agregar columnas C-F
    ws_pe[f'C{fila}'] = f"=IFERROR($B$4*(1+B{fila}),0)"
    ws_pe[f'C{fila}'].number_format = "$#,##0"

    ws_pe[f'D{fila}'] = f"=IFERROR($B$11/(C{fila}-$B$5),0)"
    ws_pe[f'D{fila}'].number_format = "#,##0"

    ws_pe[f'E{fila}'] = f"=IFERROR(D{fila}*C{fila},0)"
    ws_pe[f'E{fila}'].number_format = "$#,##0"

    ws_pe[f'F{fila}'] = f"=IFERROR((D{fila}-$B$16)/$B$16,0)"
    ws_pe[f'F{fila}'].number_format = "0.0%"

print("   ✓ Punto de Equilibrio - Análisis de sensibilidad agregado")

# ============================================================================
# PROYECCIONES - Corregir filas
# ============================================================================
print("\n📈 Corrigiendo PROYECCIONES...")
ws_proy = wb["PROYECCIONES"]

# Corregir supuestos (fila 5 = Tasa crecimiento, está como 0.1, debería ser fórmula)
ws_proy['B5'] = 0.10
ws_proy['B5'].number_format = "0.0%"

# Fila 6 ya tiene =INDICADORES!G16 ✓

ws_proy['B7'] = 0.15
ws_proy['B7'].number_format = "0.0%"

ws_proy['B8'] = 0.10
ws_proy['B8'].number_format = "0.0%"

# ESTADO DE RESULTADOS PROYECTADO
# Según la estructura: Fila 18 = Ingresos (Ventas)

# CORREGIR: Mover fórmulas a las filas correctas
# Fila 18: Ingresos (Ventas)
ws_proy['B18'] = "=IFERROR(EERR!G5*(1+$B$5),0)"
ws_proy['C18'] = "=IFERROR(B18*(1+$B$5),0)"
ws_proy['D18'] = "=IFERROR(C18*(1+$B$5),0)"
ws_proy['E18'] = "=IFERROR(D18*(1+$B$5),0)"
ws_proy['F18'] = "=IFERROR(E18*(1+$B$5),0)"
ws_proy['G18'] = "=IFERROR(F18*(1+$B$5),0)"
ws_proy['H18'] = "=IFERROR(G18*(1+$B$5),0)"

# Fila 19: Costo de Ventas
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_proy[f'{col}19'] = f"=IFERROR({col}18*(1-$B$6),0)"

# Fila 20: Utilidad Bruta
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_proy[f'{col}20'] = f"=IFERROR({col}18-{col}19,0)"

# Fila 22: Gastos Administrativos
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_proy[f'{col}22'] = f"=IFERROR({col}18*$B$7,0)"

# Fila 23: Gastos de Ventas
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_proy[f'{col}23'] = f"=IFERROR({col}18*$B$8,0)"

# Fila 24: Depreciación
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_proy[f'{col}24'] = f"=IFERROR({col}18*0.02,0)"

# Fila 25: EBITDA
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_proy[f'{col}25'] = f"=IFERROR({col}20-{col}22-{col}23,0)"

# Fila 26: Utilidad Operativa (EBIT)
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_proy[f'{col}26'] = f"=IFERROR({col}25-{col}24,0)"

# Fila 28: Gastos Financieros
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_proy[f'{col}28'] = f"=IFERROR({col}18*0.02,0)"

# Fila 29: Utilidad Antes de Impuestos
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_proy[f'{col}29'] = f"=IFERROR({col}26-{col}28,0)"

# Fila 30: Impuesto de Renta
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_proy[f'{col}30'] = f"=IFERROR({col}29*0.35,0)"

# Fila 31: UTILIDAD NETA
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_proy[f'{col}31'] = f"=IFERROR({col}29-{col}30,0)"

# Formato
for fila in range(18, 32):
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws_proy[f'{col}{fila}'].number_format = "#,##0"

print("   ✓ Proyecciones - Todas las fórmulas corregidas")

# ============================================================================
# VALORACIÓN - Corregir filas
# ============================================================================
print("\n💰 Corrigiendo VALORACIÓN...")
ws_val = wb["VALORACIÓN"]

# Las fórmulas ya están pero en filas incorrectas, déjame corregir según estructura real

# WACC ya calculado en B24 ✓

# FLUJO DE CAJA LIBRE DESCONTADO
# Fila 27: Encabezado "Año" "FCF" "Factor Desc" "FCF Descontado"
# Necesito agregar en fila 28 los años, fila 29 FCF, fila 30 factor, fila 31 FCF desc

# Fila 28: Años 1-7
for i, col in enumerate(['B', 'C', 'D', 'E', 'F', 'G', 'H'], start=1):
    ws_val[f'{col}28'] = i

# Fila 29: FCF (desde Proyecciones EBIT * 0.65)
for i, col in enumerate(['B', 'C', 'D', 'E', 'F', 'G', 'H'], start=1):
    # El EBIT proyectado está en fila 26 de PROYECCIONES
    ws_val[f'{col}29'] = f"=IFERROR(PROYECCIONES!{col}26*0.65,0)"
    ws_val[f'{col}29'].number_format = "$#,##0"

# Fila 30: Factor de Descuento
for i, col in enumerate(['B', 'C', 'D', 'E', 'F', 'G', 'H'], start=1):
    ws_val[f'{col}30'] = f"=IFERROR(1/((1+$B$24)^{i}),0)"
    ws_val[f'{col}30'].number_format = "0.0000"

# Fila 31: FCF Descontado
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_val[f'{col}31'] = f"=IFERROR({col}29*{col}30,0)"
    ws_val[f'{col}31'].number_format = "$#,##0"

# Suma FCF Descontados (fila 33)
ws_val['B33'] = "=IFERROR(SUM(B31:H31),0)"
ws_val['B33'].number_format = "$#,##0"
ws_val['B33'].font = Font(bold=True)
ws_val['A33'] = "SUMA FCF Descontados:"
ws_val['A33'].font = Font(bold=True)

# Valor Terminal (fila 35-37)
ws_val['A35'] = "Valor Terminal (Perpetuidad):"
ws_val['A35'].font = Font(bold=True)

ws_val['A36'] = "Tasa de crecimiento perpetuo:"
ws_val['B36'] = 0.025
ws_val['B36'].number_format = "0.0%"

ws_val['A37'] = "VT ="
ws_val['B37'] = "=IFERROR(H29*(1+B36)/(B24-B36),0)"
ws_val['B37'].number_format = "$#,##0"

ws_val['A38'] = "VT Descontado ="
ws_val['B38'] = "=IFERROR(B37*H30,0)"
ws_val['B38'].number_format = "$#,##0"

# Valor Total de la Empresa (fila 40)
ws_val['A40'] = "VALOR TOTAL DE LA EMPRESA:"
ws_val['A40'].font = Font(bold=True, size=14)
ws_val['B40'] = "=IFERROR(B33+B38,0)"
ws_val['B40'].number_format = "$#,##0"
ws_val['B40'].font = Font(bold=True, size=14)
ws_val['B40'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

print("   ✓ Valoración - Todas las fórmulas corregidas")

# ============================================================================
# GUARDAR
# ============================================================================
print("\n💾 Guardando archivo...")
wb.save(archivo)

print("\n" + "="*80)
print("✅ FÓRMULAS CORREGIDAS EN FILAS EXACTAS")
print("="*80)
print(f"\nArchivo actualizado: {archivo}")
print("\n📊 Hojas corregidas:")
print("   ✓ PUNTO EQUILIBRIO - Análisis de sensibilidad completo")
print("   ✓ PROYECCIONES - Estado de resultados 7 años (filas 18-31)")
print("   ✓ VALORACIÓN - FCF, Valor Terminal, Valor Total (filas 28-40)")
print("\n🎯 ¡Ahora todas las fórmulas están en las filas correctas!")
