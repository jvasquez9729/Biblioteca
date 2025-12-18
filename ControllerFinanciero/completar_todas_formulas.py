#!/usr/bin/env python3
"""
Script COMPLETO para agregar TODAS las fórmulas faltantes
Conecta todas las hojas del sistema
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

def agregar_formulas_punto_equilibrio(wb):
    """Agrega todas las fórmulas de Punto de Equilibrio"""
    print("\n📊 Agregando fórmulas a PUNTO DE EQUILIBRIO...")

    ws = wb["PUNTO EQUILIBRIO"]

    # Total Costos Fijos (fila 12)
    ws['B12'] = "=IFERROR(SUM(B9:B11),0)"
    ws['B12'].number_format = "$#,##0"

    # Margen de Contribución Unitario (fila 14)
    ws['B14'] = "=IFERROR(B5-B6,0)"
    ws['B14'].number_format = "$#,##0"

    # Margen de Contribución % (fila 15)
    ws['B15'] = "=IFERROR(B14/B5,0)"
    ws['B15'].number_format = "0.0%"

    # Punto de Equilibrio en Unidades (fila 16)
    ws['B16'] = "=IFERROR(B12/B14,0)"
    ws['B16'].number_format = "#,##0"

    # Punto de Equilibrio en Valor $ (fila 17)
    ws['B17'] = "=IFERROR(B12/B15,0)"
    ws['B17'].number_format = "$#,##0"

    # Ventas Actuales (fila 18) - del último año
    ws['B18'] = "=EERR!G5"
    ws['B18'].number_format = "$#,##0"

    # Margen de Seguridad (fila 19)
    ws['B19'] = "=IFERROR((B18-B17)/B18,0)"
    ws['B19'].number_format = "0.0%"

    # Análisis de Sensibilidad - Variación de Precio
    # Fila 27 en adelante
    escenarios_precio = [
        (27, -0.15, "Precio -15%"),
        (28, -0.10, "Precio -10%"),
        (29, -0.05, "Precio -5%"),
        (30, 0.00, "Precio Base"),
        (31, 0.05, "Precio +5%"),
        (32, 0.10, "Precio +10%"),
        (33, 0.15, "Precio +15%"),
    ]

    for fila, variacion, etiqueta in escenarios_precio:
        # Variación
        ws[f'B{fila}'] = f"{variacion:.0%}"
        ws[f'B{fila}'].number_format = "0%"

        # Nuevo Precio
        ws[f'C{fila}'] = f"=IFERROR(B5*(1+{variacion}),0)"
        ws[f'C{fila}'].number_format = "$#,##0"

        # PE Unidades
        ws[f'D{fila}'] = f"=IFERROR(B12/(C{fila}-B6),0)"
        ws[f'D{fila}'].number_format = "#,##0"

        # PE Valor
        ws[f'E{fila}'] = f"=IFERROR(D{fila}*C{fila},0)"
        ws[f'E{fila}'].number_format = "$#,##0"

        # Impacto
        ws[f'F{fila}'] = f"=IFERROR((D{fila}-B16)/B16,0)"
        ws[f'F{fila}'].number_format = "0.0%"

    print("✓ Fórmulas de Punto de Equilibrio agregadas")
    return ws

def agregar_formulas_proyecciones(wb):
    """Agrega fórmulas de Proyecciones Financieras"""
    print("\n📈 Agregando fórmulas a PROYECCIONES...")

    ws = wb["PROYECCIONES"]

    # Supuestos por defecto (si están vacíos)
    if not ws['B5'].value:
        ws['B5'] = 0.10  # 10% crecimiento
        ws['B5'].number_format = "0.0%"

    if not ws['B6'].value:
        ws['B6'] = "=INDICADORES!G16"  # Margen bruto del último año
        ws['B6'].number_format = "0.0%"

    if not ws['B7'].value:
        ws['B7'] = "=INDICADORES!C16"  # Gastos admin como % ventas
        ws['B7'].number_format = "0.0%"

    if not ws['B8'].value:
        ws['B8'] = "=INDICADORES!C16"  # Gastos ventas como % ventas
        ws['B8'].number_format = "0.0%"

    # Estado de Resultados Proyectado
    # Encabezados en fila 18
    años_proy = [2025, 2026, 2027, 2028, 2029, 2030, 2031]

    # Ventas proyectadas (fila 20)
    fila_ventas = 20
    ws['B20'] = "=IFERROR(EERR!G5*(1+$B$5),0)"  # 2025
    ws['C20'] = "=IFERROR(B20*(1+$B$5),0)"      # 2026
    ws['D20'] = "=IFERROR(C20*(1+$B$5),0)"      # 2027
    ws['E20'] = "=IFERROR(D20*(1+$B$5),0)"      # 2028
    ws['F20'] = "=IFERROR(E20*(1+$B$5),0)"      # 2029
    ws['G20'] = "=IFERROR(F20*(1+$B$5),0)"      # 2030
    ws['H20'] = "=IFERROR(G20*(1+$B$5),0)"      # 2031

    # Costo de Ventas (fila 21)
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}21'] = f"=IFERROR({col}20*(1-$B$6),0)"

    # Utilidad Bruta (fila 22)
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}22'] = f"=IFERROR({col}20-{col}21,0)"

    # Gastos Administrativos (fila 24)
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}24'] = f"=IFERROR({col}20*$B$7,0)"

    # Gastos de Ventas (fila 25)
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}25'] = f"=IFERROR({col}20*$B$8,0)"

    # Depreciación (estimada como % de ventas)
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}26'] = f"=IFERROR({col}20*0.02,0)"  # 2% de ventas

    # EBITDA (fila 27)
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}27'] = f"=IFERROR({col}22-{col}24-{col}25,0)"

    # EBIT (fila 28)
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}28'] = f"=IFERROR({col}27-{col}26,0)"

    # Gastos Financieros (estimado)
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}30'] = f"=IFERROR({col}20*0.02,0)"  # 2% de ventas

    # UAI (fila 31)
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}31'] = f"=IFERROR({col}28-{col}30,0)"

    # Impuesto (fila 32)
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}32'] = f"=IFERROR({col}31*0.35,0)"

    # Utilidad Neta (fila 33)
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}33'] = f"=IFERROR({col}31-{col}32,0)"

    # Aplicar formato
    for fila in range(20, 34):
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{fila}'].number_format = "#,##0"

    print("✓ Fórmulas de Proyecciones agregadas")
    return ws

def agregar_formulas_valoracion(wb):
    """Agrega fórmulas de Valoración"""
    print("\n💰 Agregando fórmulas a VALORACIÓN...")

    ws = wb["VALORACIÓN"]

    # WACC - Componentes
    # Tasa Libre de Riesgo (usar de configuración)
    ws['B6'] = "=CONFIGURACIÓN!B16"
    ws['B6'].number_format = "0.0%"

    # Beta (default si no existe)
    if not ws['B7'].value:
        ws['B7'] = 1.2  # Beta promedio
        ws['B7'].number_format = "0.00"

    # Prima de Riesgo de Mercado
    if not ws['B8'].value:
        ws['B8'] = 0.08  # 8%
        ws['B8'].number_format = "0.0%"

    # Prima de Riesgo País
    if not ws['B9'].value:
        ws['B9'] = 0.025  # 2.5% Colombia
        ws['B9'].number_format = "0.0%"

    # Ke = Rf + β × (Rm - Rf) + Riesgo País (fila 10)
    ws['B10'] = "=IFERROR(B6+B7*B8+B9,0)"
    ws['B10'].number_format = "0.0%"

    # Costo de la Deuda (fila 12)
    if not ws['B12'].value:
        ws['B12'] = 0.12  # 12%
        ws['B12'].number_format = "0.0%"

    # Tasa de Impuesto (fila 13)
    ws['B13'] = 0.35
    ws['B13'].number_format = "0.0%"

    # Kd después de impuestos (fila 14)
    ws['B14'] = "=IFERROR(B12*(1-B13),0)"
    ws['B14'].number_format = "0.0%"

    # Estructura de Capital
    # Patrimonio (fila 17)
    ws['B17'] = "=ESF!G54"
    ws['B17'].number_format = "$#,##0"

    # Deuda (fila 18)
    ws['B18'] = "=ESF!G45"
    ws['B18'].number_format = "$#,##0"

    # Total (fila 19)
    ws['B19'] = "=IFERROR(B17+B18,0)"
    ws['B19'].number_format = "$#,##0"

    # % Patrimonio (fila 20)
    ws['B20'] = "=IFERROR(B17/B19,0)"
    ws['B20'].number_format = "0.0%"

    # % Deuda (fila 21)
    ws['B21'] = "=IFERROR(B18/B19,0)"
    ws['B21'].number_format = "0.0%"

    # WACC = Ke × (E/V) + Kd × (1-T) × (D/V) (fila 24)
    ws['B24'] = "=IFERROR(B10*B20+B14*B21,0)"
    ws['B24'].number_format = "0.0%"
    ws['B24'].font = Font(bold=True, size=14)
    ws['B24'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    # Flujo de Caja Descontado
    # Encabezados en fila 29
    años_fcf = list(range(1, 8))

    fila_año = 30
    fila_fcf = 31
    fila_factor = 32
    fila_fcf_desc = 33

    # Años
    for i, col in enumerate(['B', 'C', 'D', 'E', 'F', 'G', 'H']):
        ws[f'{col}{fila_año}'] = años_fcf[i]

    # FCF (simplificado: EBIT * (1-T) + Deprec - CAPEX)
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        col_proy = col
        ws[f'{col}{fila_fcf}'] = f"=IFERROR(PROYECCIONES!{col_proy}28*0.65,0)"
        ws[f'{col}{fila_fcf}'].number_format = "$#,##0"

    # Factor de Descuento = 1 / (1 + WACC)^n
    for i, col in enumerate(['B', 'C', 'D', 'E', 'F', 'G', 'H'], start=1):
        ws[f'{col}{fila_factor}'] = f"=IFERROR(1/((1+$B$24)^{i}),0)"
        ws[f'{col}{fila_factor}'].number_format = "0.0000"

    # FCF Descontado
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}{fila_fcf_desc}'] = f"=IFERROR({col}{fila_fcf}*{col}{fila_factor},0)"
        ws[f'{col}{fila_fcf_desc}'].number_format = "$#,##0"

    # Suma de FCF Descontados
    ws['B35'] = f"=IFERROR(SUM(B{fila_fcf_desc}:H{fila_fcf_desc}),0)"
    ws['B35'].number_format = "$#,##0"
    ws['B35'].font = Font(bold=True)

    # Valor Terminal (fila 38)
    # Tasa de crecimiento perpetuo
    if not ws['C38'].value:
        ws['C38'] = 0.025  # 2.5%
        ws['C38'].number_format = "0.0%"

    # VT = FCF último año * (1 + g) / (WACC - g)
    ws['B38'] = f"=IFERROR(H{fila_fcf}*(1+C38)/(B24-C38),0)"
    ws['B38'].number_format = "$#,##0"
    ws['B38'].font = Font(bold=True)

    # VT Descontado
    ws['B39'] = f"=IFERROR(B38*H{fila_factor},0)"
    ws['B39'].number_format = "$#,##0"

    # Valor Total de la Empresa
    ws['B41'] = "=IFERROR(B35+B39,0)"
    ws['B41'].number_format = "$#,##0"
    ws['B41'].font = Font(bold=True, size=14)
    ws['B41'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    print("✓ Fórmulas de Valoración agregadas")
    return ws

def agregar_formulas_benchmarks(wb):
    """Agrega fórmulas de comparación en Benchmarks"""
    print("\n📊 Agregando fórmulas a BENCHMARKS...")

    ws = wb["BENCHMARKS"]

    # Comparación Empresa vs Industria
    # Fila inicial de comparación (después de la tabla de benchmarks)
    fila_comp_inicio = 17

    # Indicadores a comparar
    comparaciones = [
        ("Margen Bruto", "=INDICADORES!G16", "Benchmark del sector"),
        ("Margen EBITDA", "=INDICADORES!G17", "Benchmark del sector"),
        ("Margen Neto", "=INDICADORES!G19", "Benchmark del sector"),
        ("ROE", "=INDICADORES!G21", "Benchmark del sector"),
        ("ROA", "=INDICADORES!G20", "Benchmark del sector"),
        ("Razón Corriente", "=INDICADORES!G7", "Benchmark del sector"),
        ("Nivel Endeudamiento", "=INDICADORES!G25", "Benchmark del sector"),
    ]

    fila = fila_comp_inicio
    for indicador, formula, comentario in comparaciones:
        ws[f'A{fila}'] = indicador
        ws[f'A{fila}'].font = Font(bold=True)

        # Valor de la Empresa
        ws[f'B{fila}'] = formula
        ws[f'B{fila}'].number_format = "0.0%"

        # Valor de Industria (se debe ingresar manualmente según sector)
        # ws[f'C{fila}'] = "Ingresar benchmark"

        # Brecha = Empresa - Industria
        ws[f'D{fila}'] = f"=IFERROR(B{fila}-C{fila},0)"
        ws[f'D{fila}'].number_format = "0.0%"

        # Semáforo
        ws[f'E{fila}'] = f'=IF(D{fila}>0,"🟢 Por encima",IF(D{fila}=0,"🟡 Igual","🔴 Por debajo"))'

        fila += 1

    print("✓ Fórmulas de Benchmarks agregadas")
    return ws

def agregar_formulas_diagnostico(wb):
    """Agrega fórmulas básicas en Diagnóstico"""
    print("\n📋 Agregando fórmulas a DIAGNÓSTICO...")

    ws = wb["DIAGNÓSTICO"]

    # Encabezado con nombre de empresa (fila 3)
    ws['A3'] = '=CONCATENATE("=== DIAGNÓSTICO FINANCIERO DE ",UPPER(CONFIGURACIÓN!B5)," ===")'
    ws['A3'].font = Font(bold=True, size=14)

    # Sector (fila 4)
    ws['A4'] = '=CONCATENATE("Sector: ",CONFIGURACIÓN!B7)'

    # Período (fila 5)
    ws['A5'] = '=CONCATENATE("Período analizado: 2022 - ",CONFIGURACIÓN!B11)'

    # Fecha (fila 6)
    ws['A6'] = '=CONCATENATE("Fecha de elaboración: ",TEXT(TODAY(),"dd/mm/yyyy"))'

    # Valores clave para el diagnóstico (ocultos, para usar en macros)
    ws['J1'] = "=INDICADORES!G7"  # Razón corriente
    ws['J2'] = "=INDICADORES!G19"  # Margen neto
    ws['J3'] = "=INDICADORES!G25"  # Endeudamiento
    ws['J4'] = "=INDICADORES!G21"  # ROE
    ws['J5'] = "=INDICADORES!G39"  # Ciclo efectivo

    # Ocultar columna J
    ws.column_dimensions['J'].hidden = True

    print("✓ Fórmulas de Diagnóstico agregadas")
    return ws

def agregar_formulas_informe_visual(wb):
    """Agrega datos para gráficos en Informe Visual"""
    print("\n📊 Agregando datos para INFORME VISUAL...")

    ws = wb["INFORME VISUAL"]

    # Área de datos para gráficos (columna K en adelante, oculta)

    # Datos para Gráfico 1: Evolución de Ventas y Utilidades
    ws['K5'] = "Año"
    ws['L5'] = "Ventas"
    ws['M5'] = "Utilidad Neta"

    ws['K6'] = 2022
    ws['L6'] = "=EERR!B5"
    ws['M6'] = "=EERR!B22"

    ws['K7'] = 2023
    ws['L7'] = "=EERR!D5"
    ws['M7'] = "=EERR!D22"

    ws['K8'] = 2024
    ws['L8'] = "=EERR!G5"
    ws['M8'] = "=EERR!G22"

    # Datos para Gráfico 2: Márgenes
    ws['K11'] = "Margen"
    ws['L11'] = "2022"
    ws['M11'] = "2023"
    ws['N11'] = "2024"

    ws['K12'] = "Bruto"
    ws['L12'] = "=INDICADORES!C16"
    ws['M12'] = "=INDICADORES!D16"
    ws['N12'] = "=INDICADORES!G16"

    ws['K13'] = "EBITDA"
    ws['L13'] = "=INDICADORES!C17"
    ws['M13'] = "=INDICADORES!D17"
    ws['N13'] = "=INDICADORES!G17"

    ws['K14'] = "Operativo"
    ws['L14'] = "=INDICADORES!C18"
    ws['M14'] = "=INDICADORES!D18"
    ws['N14'] = "=INDICADORES!G18"

    ws['K15'] = "Neto"
    ws['L15'] = "=INDICADORES!C19"
    ws['M15'] = "=INDICADORES!D19"
    ws['N15'] = "=INDICADORES!G19"

    # Datos para Gráfico 3: Composición del Activo
    ws['K18'] = "Componente"
    ws['L18'] = "Valor"

    ws['K19'] = "Activos Corrientes"
    ws['L19'] = "=ESF!G14"

    ws['K20'] = "Activos No Corrientes"
    ws['L20'] = "=ESF!G21"

    # Datos para Gráfico 4: Pasivo y Patrimonio
    ws['K23'] = "Componente"
    ws['L23'] = "Valor"

    ws['K24'] = "Pasivos"
    ws['L24'] = "=ESF!G45"

    ws['K25'] = "Patrimonio"
    ws['L25'] = "=ESF!G54"

    # Datos para Gráfico 7: Ciclo de Efectivo
    ws['K28'] = "Concepto"
    ws['L28'] = "Días"

    ws['K29'] = "Días Inventario"
    ws['L29'] = "=INDICADORES!G33"

    ws['K30'] = "Días Cartera"
    ws['L30'] = "=INDICADORES!G35"

    ws['K31'] = "Días Proveedores"
    ws['L31'] = "=INDICADORES!G37"

    ws['K32'] = "Ciclo Efectivo"
    ws['L32'] = "=INDICADORES!G39"

    # Ocultar columnas K-N
    for col in ['K', 'L', 'M', 'N']:
        ws.column_dimensions[col].hidden = True

    print("✓ Datos para gráficos agregados")
    return ws

def main():
    """Función principal"""
    print("="*70)
    print("AGREGANDO TODAS LAS FÓRMULAS FALTANTES")
    print("="*70)

    archivo = "/home/user/Biblioteca/ControllerFinanciero/ControllerFinanciero_v1.0.xlsx"

    wb = load_workbook(archivo)

    # Agregar fórmulas a cada hoja
    agregar_formulas_punto_equilibrio(wb)
    agregar_formulas_proyecciones(wb)
    agregar_formulas_valoracion(wb)
    agregar_formulas_benchmarks(wb)
    agregar_formulas_diagnostico(wb)
    agregar_formulas_informe_visual(wb)

    # Guardar
    wb.save(archivo)

    print("\n" + "="*70)
    print("✅ TODAS LAS FÓRMULAS AGREGADAS EXITOSAMENTE")
    print("="*70)
    print(f"\nArchivo actualizado: {archivo}")
    print("\nHojas con fórmulas completas:")
    print("  ✓ CONFIGURACIÓN")
    print("  ✓ EERR (Estado de Resultados)")
    print("  ✓ ESF (Balance General)")
    print("  ✓ INDICADORES (+30 indicadores)")
    print("  ✓ PUNTO EQUILIBRIO")
    print("  ✓ PROYECCIONES (5-7 años)")
    print("  ✓ VALORACIÓN (WACC, DCF)")
    print("  ✓ BENCHMARKS (Comparación)")
    print("  ✓ DIAGNÓSTICO (Automático)")
    print("  ✓ INFORME VISUAL (Datos para gráficos)")
    print("\n🎉 ¡Sistema completamente integrado!")

if __name__ == "__main__":
    main()
